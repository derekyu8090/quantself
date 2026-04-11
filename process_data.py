#!/usr/bin/env python3
"""
Apple Health Data Pipeline
Processes Apple Health export XML → JSON files for the Health Dashboard.

Usage:
    python3 process_data.py /path/to/apple_health_export

Output:
    public/data/cardiovascular.json
    public/data/sleep.json
    public/data/activity.json
    public/data/overview.json
"""

import xml.etree.ElementTree as ET
import json
import os
import sys
import statistics
from datetime import datetime, timedelta
from collections import defaultdict
import traceback

# Load configuration
_CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'health_config.json')
def load_config():
    if os.path.exists(_CONFIG_PATH):
        with open(_CONFIG_PATH) as f:
            return json.load(f)
    return {}

CONFIG = load_config()


def parse_date(s):
    """Parse Apple Health date string to date only."""
    if not s:
        return None
    try:
        return datetime.strptime(s[:19], '%Y-%m-%d %H:%M:%S')
    except (ValueError, TypeError):
        return None


def score_to_level(score):
    """Map 0-100 score to risk level string."""
    if score >= 70: return 'high'
    if score >= 55: return 'medium-high'
    if score >= 40: return 'medium'
    if score >= 25: return 'low-medium'
    return 'low'


def compute_risks(rhr_list, hrv_records, vo2_list, nightly_list, steps_daily,
                  body_mass, body_fat, spo2_records, sleep_breathing,
                  workouts, exercise_time_daily, all_beds, age,
                  daylight_daily=None, headphone_exposure_daily=None,
                  walking_asymmetry_daily=None, arboleaf_data=None):
    """Compute 8 risk dimensions from actual health data. Returns dict."""

    risks = {}

    # Pre-compute shared values to avoid duplication
    rhr_values = [r['value'] for r in rhr_list]
    rhr_mean = sum(rhr_values) / len(rhr_values) if rhr_values else 0
    rhr_std = (sum((v - rhr_mean) ** 2 for v in rhr_values) / len(rhr_values)) ** 0.5 if len(rhr_values) > 1 else 0
    hrv_values = [r['value'] for r in hrv_records]

    # 1. Circadian Rhythm Risk
    # Based on: bedtime std deviation, % nights after midnight, bedtime consistency
    if all_beds:
        bed_std = statistics.stdev(all_beds) if len(all_beds) > 1 else 0
        after_midnight_pct = sum(1 for b in all_beds if b > 24) / len(all_beds) * 100
        after_2am_pct = sum(1 for b in all_beds if b > 26) / len(all_beds) * 100

        # Score: 0 (healthy) to 100 (severe disruption)
        # bed_std > 2h = very bad, after_2am > 50% = very bad
        score = min(100, int(
            (min(bed_std, 4) / 4 * 40) +       # std dev component (max 40)
            (min(after_midnight_pct, 100) / 100 * 30) +  # after midnight (max 30)
            (min(after_2am_pct, 100) / 100 * 30)         # after 2am (max 30)
        ))
        # Add daylight component to circadian risk
        if daylight_daily:
            dl_vals = [v for v in daylight_daily.values() if v > 0]
            if dl_vals:
                avg_dl = sum(dl_vals[-30:]) / len(dl_vals[-30:]) if len(dl_vals) >= 30 else sum(dl_vals) / len(dl_vals)
                if avg_dl < 15:
                    score = min(100, score + 10)  # very low daylight worsens circadian risk
                elif avg_dl > 45:
                    score = max(0, score - 5)  # good daylight is protective

        risks['circadianRhythm'] = {
            'level': score_to_level(score),
            'score': score,
            'label': '昼夜节律障碍',
        }
    else:
        risks['circadianRhythm'] = {'level': 'low', 'score': 0, 'label': '昼夜节律障碍'}

    # 2. Cardio Fitness Risk
    # Based on: VO2Max percentile for age, VO2Max trend direction
    if vo2_list:
        latest_vo2 = vo2_list[-1]['value']
        # Age-adjusted thresholds for males 25-30
        if latest_vo2 >= 49: vo2_score = 5
        elif latest_vo2 >= 44: vo2_score = 25
        elif latest_vo2 >= 39: vo2_score = 50
        elif latest_vo2 >= 35: vo2_score = 75
        else: vo2_score = 90

        # Trend: compare last 3 months vs previous 3 months
        if len(vo2_list) >= 4:
            recent = [v['value'] for v in vo2_list[-3:]]
            older = [v['value'] for v in vo2_list[:-3]]
            if older:
                trend_delta = sum(recent) / len(recent) - sum(older) / len(older)
                if trend_delta < -3: vo2_score = min(100, vo2_score + 15)   # declining fast
                elif trend_delta < -1: vo2_score = min(100, vo2_score + 5)  # declining slowly
                elif trend_delta > 1: vo2_score = max(0, vo2_score - 10)    # improving

        risks['cardioFitness'] = {
            'level': score_to_level(vo2_score),
            'score': vo2_score,
            'label': '心肺功能退化',
        }
    else:
        risks['cardioFitness'] = {'level': 'low', 'score': 0, 'label': '心肺功能退化'}

    # 3. Chronic Stress Risk
    # Based on: HRV mean vs age baseline, HRV night/day ratio, RHR elevation
    if hrv_values:
        hrv_mean = sum(hrv_values) / len(hrv_values)
        # For age 28 male, healthy HRV should be 50-80ms
        if hrv_mean < 30: hrv_score = 80
        elif hrv_mean < 40: hrv_score = 60
        elif hrv_mean < 50: hrv_score = 40
        elif hrv_mean < 60: hrv_score = 25
        else: hrv_score = 10

        # Night/day ratio check
        night_hrv = [r['value'] for r in hrv_records if r['hour'] >= 22 or r['hour'] <= 6]
        day_hrv = [r['value'] for r in hrv_records if 6 < r['hour'] < 22]
        if night_hrv and day_hrv:
            ratio = sum(night_hrv) / len(night_hrv) / (sum(day_hrv) / len(day_hrv))
            if ratio < 1.05: hrv_score = min(100, hrv_score + 15)  # No night advantage = bad

        risks['chronicStress'] = {
            'level': score_to_level(hrv_score),
            'score': hrv_score,
            'label': '过劳/慢性压力',
        }
    else:
        risks['chronicStress'] = {'level': 'low', 'score': 0, 'label': '过劳/慢性压力'}

    # 4. Metabolic Risk
    # Based on: body fat %, weight trend, exercise frequency, step count
    met_score = 0
    if body_fat:
        latest_bf = body_fat[-1]['value']
        if latest_bf > 25: met_score += 40
        elif latest_bf > 20: met_score += 20
        elif latest_bf > 15: met_score += 5

    step_vals = list(steps_daily.values()) if isinstance(steps_daily, dict) else [s['value'] for s in steps_daily]
    if step_vals:
        avg_steps = sum(step_vals) / len(step_vals)
        if avg_steps < 4000: met_score += 30
        elif avg_steps < 6000: met_score += 20
        elif avg_steps < 8000: met_score += 10

    # Weight trend (last 6 months)
    if len(body_mass) >= 2:
        recent_weight = body_mass[-1]['value']
        older_weights = [b['value'] for b in body_mass[:-1]]
        avg_older = sum(older_weights) / len(older_weights)
        if recent_weight > avg_older + 3: met_score += 15   # gaining
        elif recent_weight > avg_older + 1: met_score += 5

    # Exercise frequency
    if workouts:
        days_span = max(1, (datetime.strptime(workouts[-1].get('date', '2026-01-01'), '%Y-%m-%d') -
                            datetime.strptime(workouts[0].get('date', '2020-01-01'), '%Y-%m-%d')).days)
        weekly_freq = len(workouts) / (days_span / 7)
        if weekly_freq < 1: met_score += 15
        elif weekly_freq < 2: met_score += 5

    # Enhanced metabolic risk with Arboleaf data
    if arboleaf_data:
        latest_arb = arboleaf_data[-1] if arboleaf_data else {}
        visceral = latest_arb.get('visceralFat')
        skeletal = latest_arb.get('skeletalMuscle')

        if visceral is not None:
            # Visceral fat: <9 = normal, 10-14 = high, >=15 = very high
            if visceral >= 15: met_score += 25
            elif visceral >= 10: met_score += 15
            elif visceral >= 7: met_score += 5

        if skeletal is not None:
            # Skeletal muscle %: men target >50%
            if skeletal < 40: met_score += 15
            elif skeletal < 45: met_score += 10
            elif skeletal >= 55: met_score = max(0, met_score - 5)  # good muscle = protective

    met_score = min(100, met_score)
    risks['metabolicRisk'] = {
        'level': score_to_level(met_score),
        'score': met_score,
        'label': '代谢综合征前兆',
    }

    # 5. Cardiovascular Event Risk
    # Based on: RHR anomaly frequency, SpO2 drops, HRV stability
    cv_score = 0
    if rhr_values:
        anomaly_count = sum(1 for v in rhr_values if v > rhr_mean + 2 * rhr_std)
        anomaly_pct = anomaly_count / len(rhr_values) * 100
        if anomaly_pct > 5: cv_score += 20
        elif anomaly_pct > 2: cv_score += 10

    if spo2_records:
        spo2_vals = [r['value'] for r in spo2_records]
        below90 = sum(1 for v in spo2_vals if v < 90) / len(spo2_vals) * 100
        if below90 > 2: cv_score += 25
        elif below90 > 0.5: cv_score += 10

    # Low resting HR is protective
    if rhr_values:
        if rhr_mean < 55: cv_score = max(0, cv_score - 10)  # athletic heart is protective
        elif rhr_mean > 75: cv_score += 20

    cv_score = min(100, max(0, cv_score))
    risks['cardiovascularEvent'] = {
        'level': score_to_level(cv_score),
        'score': cv_score,
        'label': '心血管急性事件',
    }

    # 6. Sleep Apnea Risk
    # Based on: breathing disturbance index, SpO2 night drops
    apnea_score = 0
    if sleep_breathing:
        avg_bd = sum(b['value'] for b in sleep_breathing) / len(sleep_breathing)
        high_bd_pct = sum(1 for b in sleep_breathing if b['value'] >= 5) / len(sleep_breathing) * 100
        if avg_bd > 5: apnea_score += 50
        elif avg_bd > 3: apnea_score += 30
        elif avg_bd > 1.5: apnea_score += 15
        apnea_score += min(30, int(high_bd_pct * 3))

    if spo2_records:
        night_spo2 = [r['value'] for r in spo2_records if r['hour'] >= 22 or r['hour'] <= 6]
        if night_spo2:
            night_below95 = sum(1 for v in night_spo2 if v < 95) / len(night_spo2) * 100
            if night_below95 > 10: apnea_score += 20
            elif night_below95 > 5: apnea_score += 10

    apnea_score = min(100, apnea_score)
    risks['sleepApnea'] = {
        'level': score_to_level(apnea_score),
        'score': apnea_score,
        'label': '睡眠呼吸障碍',
    }

    # 7. Hearing Health Risk
    hearing_score = 0
    if headphone_exposure_daily:
        all_db = []
        for vals in headphone_exposure_daily.values():
            all_db.extend(vals)
        if all_db:
            avg_db = sum(all_db) / len(all_db)
            above_80_pct = sum(1 for v in all_db if v > 80) / len(all_db) * 100
            if avg_db > 85: hearing_score += 50
            elif avg_db > 80: hearing_score += 30
            elif avg_db > 75: hearing_score += 15
            hearing_score += min(40, int(above_80_pct * 2))
        hearing_score = min(100, hearing_score)
        risks['hearingHealth'] = {
            'level': score_to_level(hearing_score),
            'score': hearing_score,
            'label': '听力健康',
        }

    # 8. Mobility Risk (from walking asymmetry)
    mobility_score = 0
    if walking_asymmetry_daily:
        asym_vals = []
        for vals in walking_asymmetry_daily.values():
            asym_vals.extend(vals)
        if asym_vals:
            avg_asym = sum(asym_vals) / len(asym_vals)
            if avg_asym > 15: mobility_score += 50
            elif avg_asym > 10: mobility_score += 30
            elif avg_asym > 7: mobility_score += 15
        mobility_score = min(100, mobility_score)
        risks['mobilityRisk'] = {
            'level': score_to_level(mobility_score),
            'score': mobility_score,
            'label': '步态/关节风险',
        }

    return risks


def compute_goals(rhr_stats, hrv_stats, steps_stats, body_mass, nightly_list, workouts, all_beds):
    """Compute personalized goal targets based on current data."""
    goals = {}

    # Bedtime goal: move 1h earlier in 2 weeks, 2h in 4 weeks
    if all_beds:
        current_bed = (round(sum(all_beds[-30:]) / len(all_beds[-30:]), 1)
                       if len(all_beds) >= 30
                       else round(sum(all_beds) / len(all_beds), 1))
        goals['bedtime'] = {
            'current': current_bed,
            'target2w': round(max(22, current_bed - 1), 1),
            'target4w': round(max(22, current_bed - 2), 1),
            'unit': 'h', 'label': '入睡时间',
        }

    # HRV goal: +5ms in 2w, +10ms in 4w
    if hrv_stats.get('mean'):
        goals['hrv'] = {
            'current': hrv_stats['mean'],
            'target2w': round(hrv_stats['mean'] + 5, 1),
            'target4w': round(hrv_stats['mean'] + 10, 1),
            'unit': 'ms', 'label': 'HRV SDNN',
        }

    # RHR goal: maintain or lower
    if rhr_stats.get('latest'):
        goals['rhr'] = {
            'current': rhr_stats['latest'],
            'target2w': min(rhr_stats['latest'], 55),
            'target4w': min(rhr_stats['latest'], 55),
            'unit': 'bpm', 'label': '静息心率',
        }

    # Steps goal: increase toward 8000
    step_vals = [s['value'] for s in steps_stats] if isinstance(steps_stats, list) else []
    current_steps = (round(sum(step_vals[-30:]) / len(step_vals[-30:]))
                     if len(step_vals) >= 30
                     else (steps_stats.get('mean', 5000) if isinstance(steps_stats, dict) else 5000))
    goals['steps'] = {
        'current': current_steps,
        'target2w': min(max(current_steps + 1000, 7000), 10000),
        'target4w': min(max(current_steps + 2000, 8000), 12000),
        'unit': '步', 'label': '日均步数',
    }

    # Exercise frequency
    if workouts and len(workouts) >= 2:
        days_span = max(7, (datetime.strptime(workouts[-1].get('date', '2026-01-01'), '%Y-%m-%d') -
                            datetime.strptime(workouts[max(0, len(workouts) - 30)].get('date', '2020-01-01'), '%Y-%m-%d')).days)
        recent_workouts = workouts[-30:]
        current_freq = round(len(recent_workouts) / (days_span / 7), 1)
    else:
        current_freq = 1.0
    goals['exerciseFreq'] = {
        'current': current_freq,
        'target2w': max(current_freq + 1, 3),
        'target4w': max(current_freq + 2, 4),
        'unit': '次/周', 'label': '周运动次数',
    }

    # Weight
    if body_mass:
        goals['weight'] = {
            'current': body_mass[-1]['value'],
            'target2w': body_mass[-1]['value'],
            'target4w': round(max(body_mass[-1]['value'] - 1, 70), 1),
            'unit': 'kg', 'label': '体重',
        }

    return goals


def compute_health_score(rhr_daily, hrv_daily_map, nightly, steps_daily, workouts, body_mass, age, daylight_daily=None):
    """
    Compute daily health score (0-100) from weighted components:
    - RHR (15%): lower is better, scored against personal baseline
    - HRV (20%): higher is better, scored against age-adjusted range
    - Sleep (25%): duration (target 7-8h) + deep sleep %
    - Activity (15%): steps vs 8000 target + exercise minutes
    - Recovery (15%): HRV stability + RHR consistency
    - Body (10%): weight trend stability

    Returns: {
        'daily': [{'date': '2023-08-04', 'score': 72, 'rhr': 80, 'hrv': 65, 'sleep': 70, 'activity': 55, 'recovery': 75, 'body': 80}],
        'latest': 72,
        'mean': 68,
        'trend': 'improving',  # or 'declining', 'stable'
        'breakdown': {'rhr': 80, 'hrv': 65, 'sleep': 70, 'activity': 55, 'recovery': 75, 'body': 80}
    }
    """
    # Build lookup maps
    rhr_map = {r['date']: r['value'] for r in rhr_daily}
    # hrv_daily_map is already {date: [values]}
    nightly_map = {n['date']: n for n in nightly}
    # steps_daily is a defaultdict(float)

    # Get all dates that have at least RHR or HRV data
    all_dates = sorted(set(list(rhr_map.keys()) + list(hrv_daily_map.keys())))

    # Personal baselines (rolling 30-day)
    rhr_values = [r['value'] for r in rhr_daily]
    rhr_baseline = sum(rhr_values) / len(rhr_values) if rhr_values else 60

    hrv_all = []
    for vs in hrv_daily_map.values():
        hrv_all.extend(vs)
    hrv_baseline = sum(hrv_all) / len(hrv_all) if hrv_all else 50

    # Weight baseline
    weight_values = [b['value'] for b in body_mass] if body_mass else []
    weight_baseline = sum(weight_values) / len(weight_values) if weight_values else 75

    daily_scores = []

    for date in all_dates:
        components = {}

        # RHR component (15%) - lower is better
        rhr_val = rhr_map.get(date)
        if rhr_val is not None:
            # Score: 100 if rhr <= baseline-5, 0 if rhr >= baseline+15
            rhr_score = max(0, min(100, 100 - (rhr_val - (rhr_baseline - 5)) / 20 * 100))
            components['rhr'] = round(rhr_score)

        # HRV component (20%) - higher is better
        hrv_vals = hrv_daily_map.get(date, [])
        if hrv_vals:
            hrv_mean = sum(hrv_vals) / len(hrv_vals)
            # Score: 0 if hrv <= 20, 100 if hrv >= 80
            hrv_score = max(0, min(100, (hrv_mean - 20) / 60 * 100))
            components['hrv'] = round(hrv_score)

        # Sleep component (25%)
        night = nightly_map.get(date)
        if night:
            total = night.get('total', 0)
            deep = night.get('deep', 0)
            # Duration score: peak at 7.5h, drops off at <6 or >9.5
            if total <= 0:
                dur_score = 0
            elif total < 6:
                dur_score = total / 6 * 60
            elif total <= 8.5:
                dur_score = 80 + (min(total, 7.5) - 6) / 1.5 * 20
            else:
                dur_score = max(50, 100 - (total - 8.5) * 20)

            # Deep sleep bonus
            deep_pct = deep / total * 100 if total > 0 else 0
            deep_bonus = min(20, deep_pct)  # up to 20 bonus points for deep sleep

            sleep_score = min(100, dur_score * 0.8 + deep_bonus)
            components['sleep'] = round(sleep_score)

        # Activity component (15%)
        step_val = steps_daily.get(date, 0)
        step_score = min(100, step_val / 10000 * 100)
        components['activity'] = round(step_score)

        # Recovery component (15%) - based on HRV relative to personal baseline
        if hrv_vals:
            hrv_mean = sum(hrv_vals) / len(hrv_vals)
            recovery_ratio = hrv_mean / hrv_baseline if hrv_baseline > 0 else 1
            recovery_score = max(0, min(100, recovery_ratio * 70 + 15))
            components['recovery'] = round(recovery_score)

        # Body component (10%) - weight stability
        components['body'] = 70  # Default stable score, adjusted below

        # Daylight component (new, 8% weight)
        if daylight_daily is not None:
            dl_val = daylight_daily.get(date, 0)
            if dl_val > 0:
                # Score: 0-10min = 20, 10-20 = 40, 20-40 = 65, 40-60 = 80, >60 = 95
                if dl_val >= 60: dl_score = 95
                elif dl_val >= 40: dl_score = 75 + (dl_val - 40) / 20 * 20
                elif dl_val >= 20: dl_score = 55 + (dl_val - 20) / 20 * 20
                elif dl_val >= 10: dl_score = 30 + (dl_val - 10) / 10 * 25
                else: dl_score = max(10, dl_val / 10 * 30)
                components['daylight'] = round(dl_score)

        # Only compute if we have enough components
        if len(components) >= 3:
            weights = CONFIG.get('scoring', {}).get('daily', {}).get('weights', {'rhr': 13, 'hrv': 18, 'sleep': 23, 'activity': 13, 'recovery': 13, 'body': 10, 'daylight': 10})
            total_weight = sum(weights[k] for k in components if k in weights)
            if total_weight > 0:
                score = sum(components.get(k, 50) * weights.get(k, 0) for k in weights) / 100
                score = round(max(0, min(100, score)))

                daily_scores.append({
                    'date': date,
                    'score': score,
                    **{k: components.get(k, None) for k in ['rhr', 'hrv', 'sleep', 'activity', 'recovery', 'body', 'daylight']}
                })

    if not daily_scores:
        return {'daily': [], 'latest': 0, 'mean': 0, 'trend': 'stable', 'breakdown': {}}

    scores = [d['score'] for d in daily_scores]
    latest = daily_scores[-1]

    # Trend: compare last 14 days vs previous 14 days
    recent = scores[-14:] if len(scores) >= 14 else scores
    older = scores[-28:-14] if len(scores) >= 28 else scores[:len(scores)//2]
    recent_mean = sum(recent) / len(recent) if recent else 0
    older_mean = sum(older) / len(older) if older else recent_mean

    if recent_mean > older_mean + 3:
        trend = 'improving'
    elif recent_mean < older_mean - 3:
        trend = 'declining'
    else:
        trend = 'stable'

    return {
        'daily': daily_scores,
        'latest': latest['score'],
        'mean': round(sum(scores) / len(scores)),
        'trend': trend,
        'breakdown': {k: latest.get(k, 0) for k in ['rhr', 'hrv', 'sleep', 'activity', 'recovery', 'body', 'daylight']},
    }


def compute_baselines(rhr_daily, hrv_daily_map, nightly, steps_daily):
    """
    Compute 30-day rolling baselines for key metrics.
    Returns dict with current deviation from baseline for each metric.

    Output: {
        'rhr': {'mean': 54.2, 'stddev': 4.0, 'current': 52, 'deviation': -0.55, 'trend': 'stable'},
        'hrv': {...},
        'sleep': {...},
        'steps': {...},
    }
    """
    baselines = {}

    # RHR baseline
    if rhr_daily:
        rhr_vals = [r['value'] for r in rhr_daily]
        recent_30 = rhr_vals[-30:] if len(rhr_vals) >= 30 else rhr_vals
        rhr_mean = sum(recent_30) / len(recent_30)
        rhr_std = (sum((v - rhr_mean)**2 for v in recent_30) / len(recent_30))**0.5 if len(recent_30) > 1 else 0
        current = rhr_vals[-1] if rhr_vals else rhr_mean
        deviation = (current - rhr_mean) / rhr_std if rhr_std > 0 else 0

        # Trend (last 7 days vs baseline)
        last_7 = rhr_vals[-7:] if len(rhr_vals) >= 7 else rhr_vals
        last_7_mean = sum(last_7) / len(last_7)
        if last_7_mean > rhr_mean + rhr_std: trend = 'elevated'
        elif last_7_mean < rhr_mean - rhr_std: trend = 'low'
        else: trend = 'normal'

        baselines['rhr'] = {
            'mean': round(rhr_mean, 1),
            'stddev': round(rhr_std, 1),
            'current': round(current, 1),
            'deviation': round(deviation, 2),
            'trend': trend,
            'alert': abs(deviation) > 1.5,
        }

    # HRV baseline
    hrv_daily_vals = {}
    for d, vs in hrv_daily_map.items():
        hrv_daily_vals[d] = sum(vs) / len(vs)
    if hrv_daily_vals:
        sorted_dates = sorted(hrv_daily_vals.keys())
        hrv_vals = [hrv_daily_vals[d] for d in sorted_dates]
        recent_30 = hrv_vals[-30:] if len(hrv_vals) >= 30 else hrv_vals
        hrv_mean = sum(recent_30) / len(recent_30)
        hrv_std = (sum((v - hrv_mean)**2 for v in recent_30) / len(recent_30))**0.5 if len(recent_30) > 1 else 0
        current = hrv_vals[-1] if hrv_vals else hrv_mean
        deviation = (current - hrv_mean) / hrv_std if hrv_std > 0 else 0

        last_7 = hrv_vals[-7:] if len(hrv_vals) >= 7 else hrv_vals
        last_7_mean = sum(last_7) / len(last_7)
        if last_7_mean > hrv_mean + hrv_std: trend = 'elevated'
        elif last_7_mean < hrv_mean - hrv_std: trend = 'low'
        else: trend = 'normal'

        baselines['hrv'] = {
            'mean': round(hrv_mean, 1),
            'stddev': round(hrv_std, 1),
            'current': round(current, 1),
            'deviation': round(deviation, 2),
            'trend': trend,
            'alert': deviation < -1.5,  # low HRV is concerning
        }

    # Sleep baseline
    if nightly:
        sleep_vals = [n['total'] for n in nightly if n.get('total', 0) > 2]
        recent_30 = sleep_vals[-30:] if len(sleep_vals) >= 30 else sleep_vals
        sleep_mean = sum(recent_30) / len(recent_30)
        sleep_std = (sum((v - sleep_mean)**2 for v in recent_30) / len(recent_30))**0.5 if len(recent_30) > 1 else 0
        current = sleep_vals[-1] if sleep_vals else sleep_mean
        deviation = (current - sleep_mean) / sleep_std if sleep_std > 0 else 0

        baselines['sleep'] = {
            'mean': round(sleep_mean, 2),
            'stddev': round(sleep_std, 2),
            'current': round(current, 2),
            'deviation': round(deviation, 2),
            'trend': 'normal',
            'alert': deviation < -1.5,
        }

    # Steps baseline
    step_vals = sorted(steps_daily.items())
    if step_vals:
        sv = [v for _, v in step_vals if v > 100]
        recent_30 = sv[-30:] if len(sv) >= 30 else sv
        step_mean = sum(recent_30) / len(recent_30)
        step_std = (sum((v - step_mean)**2 for v in recent_30) / len(recent_30))**0.5 if len(recent_30) > 1 else 0
        current = sv[-1] if sv else step_mean
        deviation = (current - step_mean) / step_std if step_std > 0 else 0

        baselines['steps'] = {
            'mean': round(step_mean),
            'stddev': round(step_std),
            'current': round(current),
            'deviation': round(deviation, 2),
            'trend': 'normal',
            'alert': deviation < -1.5,
        }

    return baselines


def compute_longevity_score(vo2_list, rhr_list, hrv_records, hrv_daily_map, nightly_list,
                            steps_daily, workouts, body_mass, body_fat, spo2_records,
                            exercise_time_daily, age, sex, daylight_daily=None):
    """
    Evidence-based Longevity Score (0-100) computed from 8 weighted components.

    Weights based on meta-analysis effect sizes:
    - VO2Max: 23% (Kodama 2009 JAMA: 13% mortality reduction per MET)
    - Sleep Regularity: 18% (Windred 2024 Sleep: SRI > duration as predictor)
    - Activity: 18% (Paluch 2022 Lancet: 8000 steps = 45% reduction)
    - HRV: 13% (Hillebrand 2013: lowest vs highest SDNN HR=1.35)
    - Daylight: 8% (circadian rhythm entrainment via SCN)
    - Body Composition: 8% (Jayedi 2022: modest independent effect)
    - SpO2: 5% (Yan 2024: HR 0.93 per unit increase)
    - Resting HR: 5% (Aune 2017: 9% per 10bpm)

    Returns: {
        'score': 65,
        'components': {
            'vo2max': {'score': 55, 'weight': 23, 'detail': 'VO2Max 42.6, 35th percentile for age 28 male'},
            ...
        },
        'monthly': [{'month': '2023-08', 'score': 62}, ...],
        'trend': 'declining',
        'references': ['Kodama 2009 JAMA', 'Windred 2024 Sleep', ...]
    }
    """
    components = {}

    # ── 1. VO2Max (25%) ──────────────────────────────────────────
    # FRIEND Registry percentile tables (Kaminsky 2015, Mayo Clin Proc)
    # Men age brackets: 20-29, 30-39, 40-49, 50-59, 60-69, 70-79
    FRIEND_MALE = {
        20: [32.1, 40.1, 48.0, 55.2, 61.8],  # 10th, 25th, 50th, 75th, 90th
        30: [30.2, 35.9, 42.4, 49.2, 56.5],
        40: [26.8, 31.9, 37.8, 45.0, 52.1],
        50: [22.8, 27.1, 32.6, 39.7, 45.6],
        60: [19.8, 23.7, 28.2, 34.5, 40.3],
        70: [17.1, 20.4, 24.4, 30.4, 36.6],
    }
    FRIEND_FEMALE = {
        20: [23.9, 30.5, 37.6, 44.7, 51.3],
        30: [20.9, 25.3, 30.2, 36.1, 41.4],
        40: [18.8, 22.1, 26.7, 32.4, 38.4],
        50: [17.3, 19.9, 23.4, 27.6, 32.0],
        60: [14.6, 17.2, 20.0, 23.8, 27.0],
        70: [13.6, 15.6, 18.3, 20.8, 23.1],
    }

    if vo2_list:
        latest_vo2 = vo2_list[-1]['value']
        table = FRIEND_MALE if sex == 'male' else FRIEND_FEMALE
        # Find age bracket
        age_bracket = max(k for k in table.keys() if k <= max(20, min(age, 70)))
        percentiles = table[age_bracket]  # [10th, 25th, 50th, 75th, 90th]

        # Interpolate percentile
        pctile_points = [10, 25, 50, 75, 90]
        if latest_vo2 <= percentiles[0]:
            pctile = max(1, 10 * latest_vo2 / percentiles[0])
        elif latest_vo2 >= percentiles[4]:
            pctile = min(99, 90 + 10 * (latest_vo2 - percentiles[4]) / (percentiles[4] - percentiles[3]))
        else:
            pctile = 50  # fallback
            for i in range(len(percentiles) - 1):
                if percentiles[i] <= latest_vo2 <= percentiles[i+1]:
                    frac = (latest_vo2 - percentiles[i]) / (percentiles[i+1] - percentiles[i])
                    pctile = pctile_points[i] + frac * (pctile_points[i+1] - pctile_points[i])
                    break

        vo2_score = round(min(100, pctile))
        components['vo2max'] = {
            'score': vo2_score,
            'weight': 23,
            'detail': f'VO2Max {latest_vo2:.1f}, {vo2_score}th percentile (age {age}, {sex})',
        }

    # ── 2. Sleep Regularity Index (20%) ──────────────────────────
    # Simplified SRI: based on bedtime consistency (Windred 2024)
    # True SRI requires epoch-level data; we approximate from bedtime std dev
    if nightly_list:
        bedtimes = [n['bedtime'] for n in nightly_list if n.get('bedtime')]
        durations = [n['total'] for n in nightly_list if n.get('total', 0) > 2]

        if len(bedtimes) >= 14:
            recent_beds = bedtimes[-30:] if len(bedtimes) >= 30 else bedtimes
            bed_mean = sum(recent_beds) / len(recent_beds)
            bed_std = (sum((b - bed_mean)**2 for b in recent_beds) / len(recent_beds))**0.5

            # SRI approximation: std < 0.5h = excellent (score ~90), std > 3h = poor (score ~20)
            # Windred 2024: median SRI 81, IQR 73.8-86.3
            regularity_score = max(0, min(100, 100 - (bed_std - 0.3) / 3.0 * 80))

            # Duration component (U-shaped: optimal 7-8h)
            avg_dur = (sum(durations[-30:]) / len(durations[-30:])
                       if len(durations) >= 30 else sum(durations) / len(durations))
            if 7.0 <= avg_dur <= 8.5:
                dur_score = 100
            elif avg_dur < 7.0:
                dur_score = max(0, 100 - (7.0 - avg_dur) * 40)
            else:
                dur_score = max(0, 100 - (avg_dur - 8.5) * 30)

            # Combined: 60% regularity + 40% duration (regularity is more predictive per Windred)
            sleep_score = round(regularity_score * 0.6 + dur_score * 0.4)
            components['sleepRegularity'] = {
                'score': sleep_score,
                'weight': 18,
                'detail': f'Bedtime std {bed_std:.1f}h, avg duration {avg_dur:.1f}h',
            }

    # ── 3. Activity (20%) ────────────────────────────────────────
    # Steps: Paluch 2022 — plateau at 8000-10000 for age <60
    # Exercise: Arem 2015 — max benefit at 450-750 min/week
    step_vals = (list(steps_daily.values()) if isinstance(steps_daily, dict)
                 else [s['value'] for s in steps_daily])
    weekly_freq = 0.0
    if step_vals:
        recent_steps = step_vals[-30:] if len(step_vals) >= 30 else step_vals
        avg_steps = sum(recent_steps) / len(recent_steps)

        # Step score: plateau at 10000 for <60, 8000 for >=60
        target = 10000 if age < 60 else 8000
        step_score = min(100, avg_steps / target * 100)

        # Exercise frequency bonus
        exercise_score = 50  # default
        if workouts:
            recent_workouts = [
                w for w in workouts
                if w.get('date', '') >= (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
            ]
            weeks = max(1, 90 / 7)
            weekly_freq = len(recent_workouts) / weeks
            # WHO: 150 min/week moderate = target
            exercise_score = min(100, weekly_freq / 3 * 70 + 15)

        activity_score = round(step_score * 0.6 + exercise_score * 0.4)
        components['activity'] = {
            'score': activity_score,
            'weight': 18,
            'detail': (
                f'Avg {avg_steps:.0f} steps/day, {weekly_freq:.1f} workouts/week'
                if workouts else f'Avg {avg_steps:.0f} steps/day'
            ),
        }

    # ── 4. HRV (15%) ────────────────────────────────────────────
    # Age-adjusted: Umetani 1998 — SDNN declines with age
    all_hrv = [v for vs in hrv_daily_map.values() for v in vs]
    if all_hrv:
        recent_hrv = all_hrv[-300:] if len(all_hrv) >= 300 else all_hrv  # ~30 days of records
        hrv_mean = sum(recent_hrv) / len(recent_hrv)

        # Age-adjusted scoring (28yr male target: 50-80ms SDNN)
        if hrv_mean >= 80:
            hrv_score = 95
        elif hrv_mean >= 60:
            hrv_score = 75 + (hrv_mean - 60) / 20 * 20
        elif hrv_mean >= 45:
            hrv_score = 55 + (hrv_mean - 45) / 15 * 20
        elif hrv_mean >= 30:
            hrv_score = 30 + (hrv_mean - 30) / 15 * 25
        else:
            hrv_score = max(5, hrv_mean / 30 * 30)

        # Night/day ratio bonus (Task Force 1996: nighttime parasympathetic dominance)
        night_hrv_vals = [r['value'] for r in hrv_records if r['hour'] >= 22 or r['hour'] <= 6]
        day_hrv_vals = [r['value'] for r in hrv_records if 6 < r['hour'] < 22]
        ratio = None
        if night_hrv_vals and day_hrv_vals:
            ratio = (sum(night_hrv_vals) / len(night_hrv_vals)
                     / (sum(day_hrv_vals) / len(day_hrv_vals)))
            if ratio >= 1.2:
                hrv_score = min(100, hrv_score + 5)   # healthy autonomic rhythm
            elif ratio < 1.0:
                hrv_score = max(0, hrv_score - 10)    # disrupted rhythm

        hrv_score = round(hrv_score)
        components['hrv'] = {
            'score': hrv_score,
            'weight': 13,
            'detail': (
                f'SDNN mean {hrv_mean:.1f}ms, night/day ratio {ratio:.2f}'
                if ratio is not None else f'SDNN mean {hrv_mean:.1f}ms'
            ),
        }

    # ── 5. Body Composition (10%) ────────────────────────────────
    # Jayedi 2022: J-shaped curve, nadir ~25% BF
    # Barry 2014: fitness matters more than fatness
    bc_score = 60  # default
    bmi = None
    if body_fat:
        latest_bf = body_fat[-1]['value']
        # Men optimal: 12-20% (ACE classification)
        if 10 <= latest_bf <= 20:
            bc_score = 85 + (1 - abs(latest_bf - 15) / 5) * 15
        elif latest_bf < 10:
            bc_score = max(40, 85 - (10 - latest_bf) * 5)
        elif latest_bf <= 25:
            bc_score = max(40, 85 - (latest_bf - 20) * 8)
        else:
            bc_score = max(10, 45 - (latest_bf - 25) * 3)
    elif body_mass:
        # Fallback to BMI estimate (assume 175cm if no height)
        latest_weight = body_mass[-1]['value']
        bmi = latest_weight / (1.75 ** 2)
        # Lancet 2016: optimal 20-25
        if 20 <= bmi <= 25:
            bc_score = 90
        elif 18.5 <= bmi < 20:
            bc_score = 75
        elif 25 < bmi <= 27.5:
            bc_score = 70
        elif 27.5 < bmi <= 30:
            bc_score = 50
        else:
            bc_score = 30

    bc_score = round(bc_score)
    components['bodyComposition'] = {
        'score': bc_score,
        'weight': 8,
        'detail': (
            f'Body fat {body_fat[-1]["value"]:.1f}%' if body_fat
            else (f'BMI ~{bmi:.1f}' if bmi is not None else 'No data')
        ),
    }

    # ── 6. SpO2 (5%) ────────────────────────────────────────────
    # Yan 2024: nocturnal SpO2 HR 0.93 per unit; Vold 2015: <=92% HR 1.99
    if spo2_records:
        spo2_vals = [r['value'] for r in spo2_records]
        spo2_mean = sum(spo2_vals) / len(spo2_vals)
        below95_pct = sum(1 for v in spo2_vals if v < 95) / len(spo2_vals) * 100

        if spo2_mean >= 97:
            spo2_score = 95
        elif spo2_mean >= 96:
            spo2_score = 85
        elif spo2_mean >= 95:
            spo2_score = 70
        elif spo2_mean >= 93:
            spo2_score = 45
        else:
            spo2_score = 20

        # Penalize frequent desaturation
        if below95_pct > 10:
            spo2_score = max(10, spo2_score - 20)
        elif below95_pct > 5:
            spo2_score = max(20, spo2_score - 10)

        spo2_score = round(spo2_score)
        components['spo2'] = {
            'score': spo2_score,
            'weight': 5,
            'detail': f'Mean {spo2_mean:.1f}%, {below95_pct:.1f}% below 95%',
        }

    # ── Daylight (8%) ────────────────────────────────────────
    # Circadian rhythm: minimum 30 min outdoor light for SCN entrainment
    if daylight_daily:
        daylight_vals = [v for v in daylight_daily.values() if v > 0]
        if daylight_vals:
            recent_dl = daylight_vals[-30:] if len(daylight_vals) >= 30 else daylight_vals
            avg_dl = sum(recent_dl) / len(recent_dl)

            # Score: <10 min = very low (20), 10-20 = low (40), 20-40 = ok (65), 40-60 = good (80), >60 = excellent (95)
            if avg_dl >= 60: dl_score = 95
            elif avg_dl >= 40: dl_score = 75 + (avg_dl - 40) / 20 * 20
            elif avg_dl >= 20: dl_score = 55 + (avg_dl - 20) / 20 * 20
            elif avg_dl >= 10: dl_score = 30 + (avg_dl - 10) / 10 * 25
            else: dl_score = max(10, avg_dl / 10 * 30)

            dl_score = round(dl_score)
            components['daylight'] = {
                'score': dl_score,
                'weight': 8,
                'detail': f'Avg {avg_dl:.0f} min/day outdoor light',
            }

    # ── 7. Resting Heart Rate (5%) ───────────────────────────────
    # Aune 2017: 9% increase per 10bpm
    if rhr_list:
        rhr_vals = [r['value'] for r in rhr_list]
        recent_rhr = rhr_vals[-30:] if len(rhr_vals) >= 30 else rhr_vals
        rhr_mean = sum(recent_rhr) / len(recent_rhr)

        if rhr_mean <= 50:
            rhr_score = 95
        elif rhr_mean <= 55:
            rhr_score = 85
        elif rhr_mean <= 60:
            rhr_score = 75
        elif rhr_mean <= 70:
            rhr_score = 55
        elif rhr_mean <= 80:
            rhr_score = 35
        else:
            rhr_score = 15

        rhr_score = round(rhr_score)
        components['restingHR'] = {
            'score': rhr_score,
            'weight': 5,
            'detail': f'Mean {rhr_mean:.0f} bpm',
        }

    # ── Weighted composite ───────────────────────────────────────
    total_weight = sum(c['weight'] for c in components.values())
    if total_weight > 0:
        raw_score = sum(c['score'] * c['weight'] for c in components.values()) / total_weight
        final_score = round(max(0, min(100, raw_score)))
    else:
        final_score = 0

    # ── Monthly trend ────────────────────────────────────────────
    # Compute simplified monthly scores using VO2Max records as anchors
    monthly_scores = []
    vo2_monthly = {}
    if vo2_list:
        for v in vo2_list:
            m = v['date'][:7]
            vo2_monthly[m] = v['value']

    if vo2_monthly:
        table = FRIEND_MALE if sex == 'male' else FRIEND_FEMALE
        age_bracket = max(k for k in table.keys() if k <= max(20, min(age, 70)))
        percentiles = table[age_bracket]
        vo2_component_score = components.get('vo2max', {}).get('score', 50)

        for month in sorted(vo2_monthly.keys()):
            vo2_val = vo2_monthly[month]
            if vo2_val <= percentiles[0]:
                pctile = 10
            elif vo2_val >= percentiles[4]:
                pctile = 95
            elif vo2_val >= percentiles[2]:
                pctile = 50 + (vo2_val - percentiles[2]) / (percentiles[4] - percentiles[2]) * 45
            else:
                pctile = 10 + (vo2_val - percentiles[0]) / (percentiles[2] - percentiles[0]) * 40

            # Use VO2Max percentile as proxy for monthly longevity score (simplified)
            offset = final_score - vo2_component_score
            monthly_est = round(max(0, min(100, pctile + offset * 0.5)))
            monthly_scores.append({'month': month, 'score': monthly_est})

    # Trend
    if len(monthly_scores) >= 4:
        recent_3 = [m['score'] for m in monthly_scores[-3:]]
        older_3 = ([m['score'] for m in monthly_scores[-6:-3]]
                   if len(monthly_scores) >= 6 else [m['score'] for m in monthly_scores[:3]])
        r_mean = sum(recent_3) / len(recent_3)
        o_mean = sum(older_3) / len(older_3)
        if r_mean > o_mean + 3:
            trend = 'improving'
        elif r_mean < o_mean - 3:
            trend = 'declining'
        else:
            trend = 'stable'
    else:
        trend = 'stable'

    references = [
        'Kodama et al. 2009, JAMA (VO2Max meta-analysis)',
        'Mandsager et al. 2018, JAMA Network Open (CRF and mortality)',
        'Windred et al. 2024, Sleep (Sleep Regularity Index)',
        'Paluch et al. 2022, Lancet Public Health (Steps dose-response)',
        'Hillebrand et al. 2013, Europace (HRV and CVD)',
        'Jayedi et al. 2022, Int J Obes (Body fat and mortality)',
        'Yan et al. 2024, J Clin Sleep Med (Nocturnal SpO2)',
        'Aune et al. 2017 (Resting HR meta-analysis)',
        "AHA Life's Essential 8, Circulation 2022",
    ]

    return {
        'score': final_score,
        'components': components,
        'monthly': monthly_scores,
        'trend': trend,
        'references': references,
    }


def compute_correlations(rhr_daily, hrv_daily_map, nightly_list, steps_daily,
                         workouts, daylight_daily):
    """
    Compute Pearson correlations between all metric pairs over the full dataset.
    Returns list of significant correlations with natural language descriptions.
    """
    import math

    # Build aligned daily metric arrays
    # Get all dates that have at least 2 metrics
    metrics = {}

    # RHR
    rhr_map = {r['date']: r['value'] for r in rhr_daily}
    if rhr_map:
        metrics['rhr'] = rhr_map

    # HRV (daily mean)
    hrv_map = {}
    for d, vs in hrv_daily_map.items():
        if vs:
            hrv_map[d] = sum(vs) / len(vs)
    if hrv_map:
        metrics['hrv'] = hrv_map

    # Sleep duration
    sleep_map = {n['date']: n['total'] for n in nightly_list if n.get('total', 0) > 2}
    if sleep_map:
        metrics['sleep'] = sleep_map

    # Steps
    step_map = {d: v for d, v in steps_daily.items() if v > 100}
    if step_map:
        metrics['steps'] = step_map

    # Daylight
    daylight_map = {d: v for d, v in daylight_daily.items() if v > 0}
    if daylight_map:
        metrics['daylight'] = daylight_map

    # Bedtime
    bedtime_map = {n['date']: n['bedtime'] for n in nightly_list if n.get('bedtime')}
    if bedtime_map:
        metrics['bedtime'] = bedtime_map

    # Deep sleep
    deep_map = {n['date']: n['deep'] for n in nightly_list if n.get('deep', 0) > 0}
    if deep_map:
        metrics['deepSleep'] = deep_map

    # Workout next-day effect (1 if workout on prev day, 0 if not)
    workout_dates = set(w.get('date', '') for w in workouts if w.get('date'))

    def pearson(x, y):
        """Compute Pearson correlation coefficient and p-value approximation."""
        n = len(x)
        if n < 10:
            return None, None
        mx = sum(x) / n
        my = sum(y) / n
        sx = math.sqrt(sum((xi - mx)**2 for xi in x) / n)
        sy = math.sqrt(sum((yi - my)**2 for yi in y) / n)
        if sx == 0 or sy == 0:
            return None, None
        r = sum((xi - mx) * (yi - my) for xi, yi in zip(x, y)) / (n * sx * sy)
        # t-test approximation for significance
        if abs(r) >= 0.999:
            return r, 0.0
        t = r * math.sqrt((n - 2) / (1 - r * r))
        # Rough p-value (significant if |t| > 2 for n > 30)
        p_significant = abs(t) > 2.0
        return r, p_significant

    # Compute all pairs
    metric_names = list(metrics.keys())
    results = []

    labels = {
        'rhr': 'Resting Heart Rate',
        'hrv': 'HRV',
        'sleep': 'Sleep Duration',
        'steps': 'Daily Steps',
        'daylight': 'Daylight Exposure',
        'bedtime': 'Bedtime',
        'deepSleep': 'Deep Sleep',
    }

    labels_zh = {
        'rhr': '静息心率',
        'hrv': '心率变异性',
        'sleep': '睡眠时长',
        'steps': '日均步数',
        'daylight': '日光暴露',
        'bedtime': '入睡时间',
        'deepSleep': '深度睡眠',
    }

    for i in range(len(metric_names)):
        for j in range(i + 1, len(metric_names)):
            m1, m2 = metric_names[i], metric_names[j]
            # Find common dates
            common_dates = sorted(set(metrics[m1].keys()) & set(metrics[m2].keys()))
            if len(common_dates) < 30:
                continue

            x = [metrics[m1][d] for d in common_dates]
            y = [metrics[m2][d] for d in common_dates]

            r, significant = pearson(x, y)
            if r is None or not significant:
                continue
            if abs(r) < 0.1:  # too weak
                continue

            # Generate natural language description
            strength = 'strong' if abs(r) >= 0.5 else 'moderate' if abs(r) >= 0.3 else 'weak'
            direction = 'positive' if r > 0 else 'negative'

            # Human-readable description
            if r > 0:
                desc_en = f"When {labels.get(m1, m1)} increases, {labels.get(m2, m2)} tends to increase ({strength})"
                desc_zh = f"当{labels_zh.get(m1, m1)}升高时，{labels_zh.get(m2, m2)}倾向于升高（{strength}相关）"
            else:
                desc_en = f"When {labels.get(m1, m1)} increases, {labels.get(m2, m2)} tends to decrease ({strength})"
                desc_zh = f"当{labels_zh.get(m1, m1)}升高时，{labels_zh.get(m2, m2)}倾向于降低（{strength}相关）"

            results.append({
                'metric1': m1,
                'metric2': m2,
                'r': round(r, 3),
                'strength': strength,
                'direction': direction,
                'n': len(common_dates),
                'description': {'en': desc_en, 'zh': desc_zh},
            })

    # Sort by absolute correlation strength
    results.sort(key=lambda x: -abs(x['r']))

    return {
        'pairs': results,
        'metrics': metric_names,
        'labels': labels,
        'labels_zh': labels_zh,
    }


def call_llm(prompt, config=None):
    """Call LLM with multi-provider support. Returns response text or raises."""
    import urllib.request
    import json as json_mod

    cfg = (config or CONFIG).get('llm', {})
    provider = cfg.get('provider', 'claude')
    providers = cfg.get('providers', {})
    pcfg = providers.get(provider, {})

    model = pcfg.get('model', 'claude-sonnet-4-20250514')
    api_url = pcfg.get('apiUrl', 'https://api.anthropic.com/v1/messages')
    env_key = pcfg.get('envKey', 'ANTHROPIC_API_KEY')
    api_key = os.environ.get(env_key, '') if env_key else ''

    # Load from .env if not in environment
    if not api_key and env_key:
        env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
        if os.path.exists(env_path):
            for line in open(env_path):
                if line.strip().startswith(env_key + '='):
                    api_key = line.strip().split('=', 1)[1].strip().strip('"\'')

    if provider == 'claude':
        req = urllib.request.Request(api_url,
            data=json_mod.dumps({
                'model': model, 'max_tokens': 1024,
                'messages': [{'role': 'user', 'content': prompt}],
            }).encode(),
            headers={
                'Content-Type': 'application/json',
                'x-api-key': api_key,
                'anthropic-version': '2023-06-01',
            })
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json_mod.loads(resp.read().decode())
        return result['content'][0]['text']

    elif provider == 'ollama':
        req = urllib.request.Request(api_url,
            data=json_mod.dumps({
                'model': model, 'stream': False,
                'messages': [{'role': 'user', 'content': prompt}],
            }).encode(),
            headers={'Content-Type': 'application/json'})
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = json_mod.loads(resp.read().decode())
        return result.get('message', {}).get('content', '')

    elif provider == 'openai':
        req = urllib.request.Request(api_url,
            data=json_mod.dumps({
                'model': model, 'max_tokens': 1024,
                'messages': [{'role': 'user', 'content': prompt}],
            }).encode(),
            headers={
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {api_key}',
            })
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json_mod.loads(resp.read().decode())
        return result['choices'][0]['message']['content']

    elif provider == 'gemini':
        url = api_url.replace('{model}', model) + f'?key={api_key}'
        req = urllib.request.Request(url,
            data=json_mod.dumps({
                'contents': [{'parts': [{'text': prompt}]}],
            }).encode(),
            headers={'Content-Type': 'application/json'})
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json_mod.loads(resp.read().decode())
        return result['candidates'][0]['content']['parts'][0]['text']

    else:
        raise ValueError(f'Unknown LLM provider: {provider}')


def generate_weekly_report(health_score, longevity_score, risks, baselines, correlations,
                           rhr_list, hrv_daily_map, nightly_list, steps_daily,
                           daylight_daily, workouts, arboleaf_data):
    """Generate a weekly health summary using LLM (multi-provider)."""

    # Build context from last 7 days
    from datetime import datetime, timedelta
    today = datetime.now().strftime('%Y-%m-%d')
    week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    two_weeks_ago = (datetime.now() - timedelta(days=14)).strftime('%Y-%m-%d')

    # Last 7 days health scores
    recent_scores = [d for d in health_score.get('daily', []) if d['date'] >= week_ago and d['date'] < today]
    prev_scores = [d for d in health_score.get('daily', []) if d['date'] >= two_weeks_ago and d['date'] < week_ago]

    # Last 7 days key metrics
    recent_rhr = [r for r in rhr_list if r['date'] >= week_ago and r['date'] < today]

    recent_hrv = []
    for d, vs in hrv_daily_map.items():
        if d >= week_ago and d < today and vs:
            recent_hrv.append({'date': d, 'value': round(sum(vs)/len(vs), 1)})

    recent_sleep = [n for n in nightly_list if n['date'] >= week_ago and n['date'] < today]

    recent_steps = {d: v for d, v in steps_daily.items() if d >= week_ago and d < today}

    recent_daylight = {d: v for d, v in daylight_daily.items() if d >= week_ago and d < today}

    recent_workouts = [w for w in workouts if w.get('date', '') >= week_ago and w.get('date', '') < today]

    # Build stats summary
    def avg(lst): return round(sum(lst)/len(lst), 1) if lst else 0

    summary = {
        'period': f'{week_ago} to {today}',
        'healthScore': {
            'thisWeek': avg([s['score'] for s in recent_scores]),
            'lastWeek': avg([s['score'] for s in prev_scores]),
        },
        'longevityScore': longevity_score.get('score', 0),
        'rhr': {
            'thisWeek': avg([r['value'] for r in recent_rhr]),
            'baseline': baselines.get('rhr', {}).get('mean', 0),
        },
        'hrv': {
            'thisWeek': avg([r['value'] for r in recent_hrv]),
            'baseline': baselines.get('hrv', {}).get('mean', 0),
        },
        'sleep': {
            'avgDuration': avg([n['total'] for n in recent_sleep]),
            'avgBedtime': avg([n.get('bedtime', 0) for n in recent_sleep if n.get('bedtime')]),
        },
        'steps': {
            'avgDaily': avg(list(recent_steps.values())) if recent_steps else 0,
        },
        'daylight': {
            'avgMinutes': avg(list(recent_daylight.values())) if recent_daylight else 0,
        },
        'workouts': len(recent_workouts),
        'workoutTypes': [w.get('type', '') for w in recent_workouts],
        'topRisks': [(v['label'], v['score'], v['level']) for k, v in sorted(risks.items(), key=lambda x: -x[1]['score'])[:3]],
        'topCorrelation': correlations.get('pairs', [{}])[0] if correlations.get('pairs') else {},
    }

    # Build prompt
    prompt = f"""You are a health data analyst. Generate a concise weekly health report in both English and Chinese.

Data for the week {summary['period']}:

- Daily Health Score: this week avg {summary['healthScore']['thisWeek']}, last week avg {summary['healthScore']['lastWeek']}
- Longevity Score: {summary['longevityScore']}/100
- Resting Heart Rate: {summary['rhr']['thisWeek']} bpm (baseline: {summary['rhr']['baseline']})
- HRV (SDNN): {summary['hrv']['thisWeek']} ms (baseline: {summary['hrv']['baseline']})
- Sleep: avg {summary['sleep']['avgDuration']:.1f}h, avg bedtime ~{summary['sleep']['avgBedtime']:.1f}h (24=midnight, 27=3am)
- Steps: avg {summary['steps']['avgDaily']:.0f}/day
- Daylight: avg {summary['daylight']['avgMinutes']:.0f} min/day
- Workouts: {summary['workouts']} sessions ({', '.join(summary['workoutTypes']) if summary['workoutTypes'] else 'none'})
- Top risks: {', '.join(f'{label} ({score})' for label, score, level in summary['topRisks'])}

Write a 150-word report with these sections:
1. **This Week** — one sentence summary (improving/declining/stable, key highlight)
2. **Key Changes** — 2-3 bullet points on what changed vs baseline or last week
3. **Action Item** — one specific, actionable recommendation for next week

Output format:
---EN---
(English report)
---ZH---
(Chinese report)"""

    # Call LLM (multi-provider)
    try:
        text = call_llm(prompt, CONFIG)

        # Parse EN and ZH sections
        en_report = ''
        zh_report = ''
        if '---EN---' in text and '---ZH---' in text:
            parts = text.split('---ZH---')
            en_report = parts[0].replace('---EN---', '').strip()
            zh_report = parts[1].strip()
        else:
            en_report = text
            zh_report = text

        return {
            'en': en_report,
            'zh': zh_report,
            'generatedAt': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'period': summary['period'],
            'error': None,
        }

    except Exception as e:
        return {
            'en': None,
            'zh': None,
            'error': str(e),
            'period': summary['period'],
        }


def compute_hr_cycles(rhr_daily, hrv_daily_map, workouts):
    """
    Detect heart rate recovery cycles (typically 3-4 weeks).
    Uses rolling mean of RHR and HRV to find periodic dips (recovery) and peaks (fatigue).
    Returns cycle data for visualization and deload recommendations.
    """
    import math

    if len(rhr_daily) < 42:  # need at least 6 weeks
        return {'cycles': [], 'avgCycleLength': 0, 'recommendation': None}

    # Build daily RHR series (last 120 days)
    rhr_map = {r['date']: r['value'] for r in rhr_daily}
    dates = sorted(rhr_map.keys())[-120:]

    # 7-day rolling mean to smooth noise
    rhr_vals = [rhr_map.get(d, 0) for d in dates]
    window = 7
    smoothed = []
    for i in range(len(rhr_vals)):
        start = max(0, i - window + 1)
        chunk = [v for v in rhr_vals[start:i+1] if v > 0]
        smoothed.append(sum(chunk) / len(chunk) if chunk else 0)

    # Find local minima (recovery troughs) and maxima (fatigue peaks)
    # A local minimum is a point lower than both 5-day neighbors
    troughs = []
    peaks = []
    margin = 5
    for i in range(margin, len(smoothed) - margin):
        if smoothed[i] <= min(smoothed[i-margin:i]) and smoothed[i] <= min(smoothed[i+1:i+margin+1]):
            troughs.append({'date': dates[i], 'value': round(smoothed[i], 1), 'day': i, 'type': 'trough'})
        if smoothed[i] >= max(smoothed[i-margin:i]) and smoothed[i] >= max(smoothed[i+1:i+margin+1]):
            peaks.append({'date': dates[i], 'value': round(smoothed[i], 1), 'day': i, 'type': 'peak'})

    # Calculate cycle lengths (trough to trough)
    cycle_lengths = []
    for i in range(1, len(troughs)):
        gap = troughs[i]['day'] - troughs[i-1]['day']
        if 14 <= gap <= 42:  # valid cycle: 2-6 weeks
            cycle_lengths.append(gap)

    avg_cycle = round(sum(cycle_lengths) / len(cycle_lengths)) if cycle_lengths else 0

    # Build smoothed series for charting
    chart_data = [{'date': dates[i], 'rhr7d': round(smoothed[i], 1)} for i in range(len(dates)) if smoothed[i] > 0]

    # Deload recommendation
    recommendation = None
    if avg_cycle > 0 and troughs:
        last_trough_day = troughs[-1]['day']
        days_since = len(dates) - 1 - last_trough_day
        days_until_next = max(0, avg_cycle - days_since)
        if days_until_next <= 5:
            recommendation = {
                'en': f'Recovery dip expected in ~{days_until_next} days. Consider a deload week.',
                'zh': f'预计 ~{days_until_next} 天后进入恢复低谷期，建议安排减量周。',
            }
        elif days_since > avg_cycle * 0.7:
            recommendation = {
                'en': f'You are {days_since} days into a ~{avg_cycle}-day cycle. Fatigue may be building.',
                'zh': f'当前处于 ~{avg_cycle} 天周期的第 {days_since} 天，疲劳可能正在积累。',
            }

    # Workout load overlay: count workouts per week for correlation
    wo_weekly = {}
    for w in workouts:
        d = w.get('date', '')
        if d:
            week = d[:7] + '-W' + str((int(d[8:10]) - 1) // 7 + 1)
            wo_weekly[week] = wo_weekly.get(week, 0) + 1

    return {
        'smoothed': chart_data,
        'troughs': troughs,
        'peaks': peaks,
        'cycleLengths': cycle_lengths,
        'avgCycleLength': avg_cycle,
        'recommendation': recommendation,
    }


def compute_trends(rhr_daily, hrv_daily_map, nightly_list, steps_daily, config=None):
    """Detect sustained trends (2+ weeks) in key metrics. Returns list of trend alerts."""
    cfg = (config or {}).get('trends', {})
    window = cfg.get('windowDays', 14)
    delta_threshold = cfg.get('significantDelta', 3)

    trends = []

    def check_trend(values, metric_name, label_en, label_zh, unit, higher_is_better=True):
        if len(values) < window * 2:
            return
        recent = values[-window:]
        previous = values[-window*2:-window]
        recent_mean = sum(recent) / len(recent)
        prev_mean = sum(previous) / len(previous)
        pct_change = ((recent_mean - prev_mean) / prev_mean * 100) if prev_mean != 0 else 0

        if abs(pct_change) < delta_threshold:
            return

        direction = 'up' if recent_mean > prev_mean else 'down'
        is_concerning = (direction == 'down' and higher_is_better) or (direction == 'up' and not higher_is_better)

        trends.append({
            'metric': metric_name,
            'direction': direction,
            'pctChange': round(pct_change, 1),
            'recentMean': round(recent_mean, 1),
            'previousMean': round(prev_mean, 1),
            'concerning': is_concerning,
            'label': {'en': label_en, 'zh': label_zh},
            'unit': unit,
            'window': window,
        })

    # RHR (lower is better)
    rhr_vals = [r['value'] for r in rhr_daily]
    check_trend(rhr_vals, 'rhr', 'Resting Heart Rate', '静息心率', 'bpm', higher_is_better=False)

    # HRV (higher is better)
    hrv_sorted = sorted(hrv_daily_map.items())
    if hrv_sorted:
        hrv_vals = [sum(vs)/len(vs) for _, vs in hrv_sorted]
        check_trend(hrv_vals, 'hrv', 'HRV', '心率变异性', 'ms', higher_is_better=True)

    # Sleep (higher is better, up to a point)
    sleep_vals = [n['total'] for n in nightly_list if n.get('total', 0) > 2]
    check_trend(sleep_vals, 'sleep', 'Sleep Duration', '睡眠时长', 'h', higher_is_better=True)

    # Steps (higher is better)
    step_items = sorted(steps_daily.items())
    if step_items:
        step_vals = [v for _, v in step_items if v > 100]
        check_trend(step_vals, 'steps', 'Daily Steps', '日均步数', 'steps', higher_is_better=True)

    return trends


def main(export_dir, arboleaf_path=None):
    import zipfile as _zipfile
    import tempfile as _tempfile

    # Support direct ZIP import — auto-extract to temp dir
    _temp_dir = None
    if export_dir.endswith('.zip') and os.path.isfile(export_dir):
        print(f"Extracting ZIP: {export_dir} ...")
        _temp_dir = _tempfile.mkdtemp(prefix='quantself_')
        with _zipfile.ZipFile(export_dir, 'r') as zf:
            zf.extractall(_temp_dir)
        # Find export.xml inside (may be nested in apple_health_export/)
        for root, dirs, files in os.walk(_temp_dir):
            if 'export.xml' in files:
                export_dir = root
                break
        else:
            print(f"ERROR: No export.xml found inside {export_dir}")
            sys.exit(1)
        print(f"  Extracted to: {export_dir}")

    xml_path = os.path.join(export_dir, 'export.xml')
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'public', 'data')
    os.makedirs(out_dir, exist_ok=True)

    # Verify XML export completeness
    if not os.path.exists(xml_path):
        print(f"ERROR: {xml_path} does not exist")
        sys.exit(1)
    file_size = os.path.getsize(xml_path)
    if file_size < 1000:
        print(f"WARNING: XML file is very small ({file_size} bytes), may be incomplete")
    else:
        with open(xml_path, 'rb') as f:
            f.seek(max(0, file_size - 200))
            tail = f.read().decode('utf-8', errors='ignore')
            if '</HealthData>' not in tail:
                print("WARNING: XML file may be truncated (missing closing </HealthData> tag)")

    print(f"Parsing {xml_path} ({file_size / 1024 / 1024:.1f} MB) ...")

    # Collectors
    rhr_daily = {}
    hrv_records = []
    vo2max_records = []
    walking_hr_records = []
    sleep_records = []
    steps_daily = defaultdict(float)
    active_energy_daily = defaultdict(float)
    exercise_time_daily = defaultdict(float)
    body_mass = []
    body_fat = []
    spo2_records = []
    resp_records = []
    sleep_breathing = []
    sleep_temp = []
    workouts = []
    activity_summaries = []
    user_info = {}
    hr_by_hour = defaultdict(list)

    walking_speed_daily = defaultdict(list)       # WalkingSpeed: m/s
    walking_step_length_daily = defaultdict(list) # WalkingStepLength: cm
    walking_asymmetry_daily = defaultdict(list)   # WalkingAsymmetryPercentage: %
    walking_steadiness = []                       # AppleWalkingSteadiness: % (sparse)
    daylight_daily = defaultdict(float)           # TimeInDaylight: minutes per day (sum)
    distance_daily = defaultdict(float)           # DistanceWalkingRunning: km per day (sum)
    flights_daily = defaultdict(float)            # FlightsClimbed: count per day (sum)
    basal_energy_daily = defaultdict(float)       # BasalEnergyBurned: kcal per day (sum)
    headphone_exposure_daily = defaultdict(list)  # HeadphoneAudioExposure: dB
    cycling_distance_daily = defaultdict(float)   # DistanceCycling: km per day (sum)
    six_min_walk = []                             # SixMinuteWalkTestDistance: meters

    record_count = 0
    source_stats = defaultdict(lambda: defaultdict(int))  # {sourceName: {recordType: count}}

    for event, elem in ET.iterparse(xml_path, events=('end',)):
        if elem.tag == 'Me':
            user_info = {
                'dob': elem.get('HKCharacteristicTypeIdentifierDateOfBirth', ''),
                'sex': elem.get('HKCharacteristicTypeIdentifierBiologicalSex', ''),
            }

        elif elem.tag == 'Record':
            rtype = elem.get('type', '')
            val_str = elem.get('value', '')
            start = elem.get('startDate', '')
            source_name = elem.get('sourceName', 'Unknown')
            dt = parse_date(start)
            if dt is None:
                elem.clear()
                continue

            date_str = dt.strftime('%Y-%m-%d')
            hour = dt.hour

            # Track source statistics
            short_type = rtype.replace('HKQuantityTypeIdentifier', '').replace('HKCategoryTypeIdentifier', '')
            source_stats[source_name][short_type] += 1

            if rtype == 'HKQuantityTypeIdentifierRestingHeartRate':
                try:
                    v = float(val_str)
                    rhr_daily[date_str] = v
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierHeartRateVariabilitySDNN':
                try:
                    v = float(val_str)
                    hrv_records.append({'date': date_str, 'hour': hour, 'value': v})
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierVO2Max':
                try:
                    v = float(val_str)
                    vo2max_records.append({'date': date_str, 'value': v})
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierWalkingHeartRateAverage':
                try:
                    v = float(val_str)
                    walking_hr_records.append({'date': date_str, 'value': v})
                except (ValueError, TypeError): pass

            elif rtype == 'HKCategoryTypeIdentifierSleepAnalysis':
                source = elem.get('sourceName', '')
                if 'Watch' in source or 'Apple' in source:
                    end = elem.get('endDate', '')
                    end_dt = parse_date(end)
                    if end_dt:
                        dur_hrs = (end_dt - dt).total_seconds() / 3600
                        sleep_records.append({
                            'start': start[:19],
                            'end': end[:19],
                            'value': val_str,
                            'duration_hrs': dur_hrs,
                            'end_date': end_dt.strftime('%Y-%m-%d'),
                            'start_hour': hour,
                        })

            elif rtype == 'HKQuantityTypeIdentifierStepCount':
                try:
                    steps_daily[date_str] += float(val_str)
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierActiveEnergyBurned':
                try:
                    active_energy_daily[date_str] += float(val_str)
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierAppleExerciseTime':
                try:
                    exercise_time_daily[date_str] += float(val_str)
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierBodyMass':
                try:
                    body_mass.append({'date': date_str, 'value': float(val_str)})
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierBodyFatPercentage':
                try:
                    body_fat.append({'date': date_str, 'value': round(float(val_str) * 100, 1)})
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierOxygenSaturation':
                try:
                    v = float(val_str) * 100
                    spo2_records.append({'date': date_str, 'hour': hour, 'value': v})
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierHeartRate':
                try:
                    hr_by_hour[hour].append(float(val_str))
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierRespiratoryRate':
                try:
                    resp_records.append({'date': date_str, 'hour': hour, 'value': float(val_str)})
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierAppleSleepingBreathingDisturbances':
                try:
                    sleep_breathing.append({'date': date_str, 'value': float(val_str)})
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierAppleSleepingWristTemperature':
                try:
                    sleep_temp.append({'date': date_str, 'value': float(val_str)})
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierWalkingSpeed':
                try:
                    walking_speed_daily[date_str].append(float(val_str))
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierWalkingStepLength':
                try:
                    walking_step_length_daily[date_str].append(float(val_str) * 100)  # m to cm
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierWalkingAsymmetryPercentage':
                try:
                    walking_asymmetry_daily[date_str].append(float(val_str) * 100)  # fraction to %
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierAppleWalkingSteadiness':
                try:
                    walking_steadiness.append({'date': date_str, 'value': round(float(val_str) * 100, 1)})
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierTimeInDaylight':
                try:
                    daylight_daily[date_str] += float(val_str)  # minutes
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierDistanceWalkingRunning':
                try:
                    distance_daily[date_str] += float(val_str)  # km
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierFlightsClimbed':
                try:
                    flights_daily[date_str] += float(val_str)
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierBasalEnergyBurned':
                try:
                    basal_energy_daily[date_str] += float(val_str)
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierHeadphoneAudioExposure':
                try:
                    headphone_exposure_daily[date_str].append(float(val_str))
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierDistanceCycling':
                try:
                    cycling_distance_daily[date_str] += float(val_str)
                except (ValueError, TypeError): pass

            elif rtype == 'HKQuantityTypeIdentifierSixMinuteWalkTestDistance':
                try:
                    six_min_walk.append({'date': date_str, 'value': round(float(val_str))})
                except (ValueError, TypeError): pass

            record_count += 1
            if record_count % 1000000 == 0:
                print(f"  Processed {record_count // 1000000}M records...")
            elem.clear()

        elif elem.tag == 'Workout':
            wo = {
                'type': elem.get('workoutActivityType', '').replace('HKWorkoutActivityType', ''),
                'startDate': elem.get('startDate', '')[:19],
                'endDate': elem.get('endDate', '')[:19],
                'duration': 0,
                'distance': 0,
                'calories': 0,
                'avgHR': 0,
                'maxHR': 0,
                'minHR': 0,
                'swimDistance': 0,
                'strokeCount': 0,
            }
            try:
                wo['duration'] = round(float(elem.get('duration', 0)))
            except (ValueError, TypeError): pass
            try:
                wo['calories'] = round(float(elem.get('totalEnergyBurned', 0)))
            except (ValueError, TypeError): pass

            for stat in elem.findall('WorkoutStatistics'):
                st = stat.get('type', '')
                if 'HeartRate' in st and 'Recovery' not in st and 'Variability' not in st and 'Walking' not in st:
                    try: wo['avgHR'] = round(float(stat.get('average', 0)))
                    except (ValueError, TypeError): pass
                    try: wo['maxHR'] = round(float(stat.get('maximum', 0)))
                    except (ValueError, TypeError): pass
                    try: wo['minHR'] = round(float(stat.get('minimum', 0)))
                    except (ValueError, TypeError): pass
                elif 'ActiveEnergyBurned' in st:
                    try: wo['calories'] = round(float(stat.get('sum', 0)))
                    except (ValueError, TypeError): pass
                elif 'DistanceSwimming' in st:
                    try: wo['swimDistance'] = round(float(stat.get('sum', 0)))
                    except (ValueError, TypeError): pass
                elif 'SwimmingStrokeCount' in st:
                    try: wo['strokeCount'] = round(float(stat.get('sum', 0)))
                    except (ValueError, TypeError): pass
                elif 'Distance' in st and 'Swimming' not in st:
                    try: wo['distance'] = round(float(stat.get('sum', 0)))
                    except (ValueError, TypeError): pass

            dt = parse_date(wo['startDate'])
            if dt:
                wo['date'] = dt.strftime('%Y-%m-%d')
                wo['hour'] = dt.hour
                workouts.append(wo)
            elem.clear()

        elif elem.tag == 'ActivitySummary':
            try:
                activity_summaries.append({
                    'date': elem.get('dateComponents', ''),
                    'activeEnergy': round(float(elem.get('activeEnergyBurned', 0))),
                    'exerciseTime': round(float(elem.get('appleExerciseTime', 0))),
                    'standHours': round(float(elem.get('appleStandHours', 0))),
                })
            except (ValueError, TypeError): pass
            elem.clear()

    print(f"  Total records: {record_count}")

    # ================================================================
    # BUILD CARDIOVASCULAR JSON
    # ================================================================
    print("Building cardiovascular.json ...")

    # RHR
    rhr_list = sorted([{'date': d, 'value': v} for d, v in rhr_daily.items()], key=lambda x: x['date'])
    rhr_values = [r['value'] for r in rhr_list]
    rhr_stats = {
        'mean': round(sum(rhr_values) / len(rhr_values), 1) if rhr_values else 0,
        'min': min(rhr_values) if rhr_values else 0,
        'max': max(rhr_values) if rhr_values else 0,
        'latest': rhr_list[-1]['value'] if rhr_list else 0,
        'latestDate': rhr_list[-1]['date'] if rhr_list else '',
    }

    # RHR monthly
    rhr_monthly = defaultdict(list)
    for r in rhr_list:
        m = r['date'][:7]
        rhr_monthly[m].append(r['value'])
    rhr_monthly_list = sorted([{
        'month': m,
        'mean': round(sum(vs)/len(vs), 1),
        'min': min(vs),
        'max': max(vs),
    } for m, vs in rhr_monthly.items()], key=lambda x: x['month'])

    # HRV
    hrv_daily_map = defaultdict(list)
    hrv_night_map = defaultdict(list)
    hrv_day_map = defaultdict(list)
    for r in hrv_records:
        hrv_daily_map[r['date']].append(r['value'])
        if r['hour'] >= 22 or r['hour'] <= 6:
            hrv_night_map[r['date']].append(r['value'])
        else:
            hrv_day_map[r['date']].append(r['value'])

    hrv_daily_list = sorted([{
        'date': d,
        'value': round(sum(vs)/len(vs), 1),
    } for d, vs in hrv_daily_map.items()], key=lambda x: x['date'])

    all_hrv = [r['value'] for r in hrv_records]
    night_hrv = [r['value'] for r in hrv_records if r['hour'] >= 22 or r['hour'] <= 6]
    day_hrv = [r['value'] for r in hrv_records if 6 < r['hour'] < 22]
    hrv_stats = {
        'mean': round(sum(all_hrv)/len(all_hrv), 1) if all_hrv else 0,
        'nightMean': round(sum(night_hrv)/len(night_hrv), 1) if night_hrv else 0,
        'dayMean': round(sum(day_hrv)/len(day_hrv), 1) if day_hrv else 0,
        'latest': hrv_daily_list[-1]['value'] if hrv_daily_list else 0,
        'latestDate': hrv_daily_list[-1]['date'] if hrv_daily_list else '',
    }

    # HRV monthly
    hrv_monthly = defaultdict(list)
    for r in hrv_records:
        m = r['date'][:7]
        hrv_monthly[m].append(r['value'])
    hrv_monthly_list = sorted([{
        'month': m,
        'mean': round(sum(vs)/len(vs), 1),
    } for m, vs in hrv_monthly.items()], key=lambda x: x['month'])

    # VO2Max
    vo2_list = sorted(vo2max_records, key=lambda x: x['date'])
    for v in vo2_list:
        v['value'] = round(v['value'], 1)
    vo2_stats = {
        'latest': vo2_list[-1]['value'] if vo2_list else 0,
        'latestDate': vo2_list[-1]['date'] if vo2_list else '',
        'peak': max(v['value'] for v in vo2_list) if vo2_list else 0,
        'mean': round(sum(v['value'] for v in vo2_list)/len(vo2_list), 1) if vo2_list else 0,
    }

    # Walking HR monthly
    whr_monthly = defaultdict(list)
    for r in walking_hr_records:
        m = r['date'][:7]
        whr_monthly[m].append(r['value'])
    whr_monthly_list = sorted([{
        'month': m,
        'mean': round(sum(vs)/len(vs), 1),
    } for m, vs in whr_monthly.items()], key=lambda x: x['month'])

    # Heart rate by hour
    hr_hourly = []
    for h in range(24):
        vals = hr_by_hour.get(h, [])
        if vals:
            hr_hourly.append({
                'hour': h,
                'mean': round(sum(vals)/len(vals)),
                'median': round(sorted(vals)[len(vals)//2]),
            })

    # SpO2 stats
    spo2_values = [r['value'] for r in spo2_records]
    spo2_stats = {
        'mean': round(sum(spo2_values)/len(spo2_values), 1) if spo2_values else 0,
        'below95pct': round(sum(1 for v in spo2_values if v < 95) / len(spo2_values) * 100, 1) if spo2_values else 0,
    }

    # SpO2 hourly profile
    spo2_by_hour = defaultdict(list)
    for r in spo2_records:
        spo2_by_hour[r['hour']].append(r['value'])
    spo2_hourly = sorted([{
        'hour': h,
        'mean': round(sum(vs)/len(vs), 1),
        'min': round(min(vs), 1),
    } for h, vs in spo2_by_hour.items()], key=lambda x: x['hour'])

    # Respiratory monthly
    resp_monthly = defaultdict(list)
    for r in resp_records:
        m = r['date'][:7]
        resp_monthly[m].append(r['value'])
    resp_monthly_list = sorted([{
        'month': m,
        'mean': round(sum(vs)/len(vs), 1),
    } for m, vs in resp_monthly.items()], key=lambda x: x['month'])

    # Respiratory nightly average
    night_resp = [r['value'] for r in resp_records if r['hour'] >= 22 or r['hour'] <= 6]
    resp_stats = {
        'mean': round(sum(r['value'] for r in resp_records) / len(resp_records), 1) if resp_records else 0,
        'nightMean': round(sum(night_resp)/len(night_resp), 1) if night_resp else 0,
    }

    cardio = {
        'rhr': {'daily': rhr_list, 'monthly': rhr_monthly_list, 'stats': rhr_stats},
        'hrv': {'daily': hrv_daily_list, 'monthly': hrv_monthly_list, 'stats': hrv_stats},
        'vo2max': {'records': vo2_list, 'stats': vo2_stats},
        'walkingHR': {'monthly': whr_monthly_list},
        'hrHourly': hr_hourly,
        'spo2': {'stats': spo2_stats, 'hourly': spo2_hourly},
        'respiratory': {'monthly': resp_monthly_list, 'stats': resp_stats},
    }

    with open(os.path.join(out_dir, 'cardiovascular.json'), 'w') as f:
        json.dump(cardio, f)
    print("  cardiovascular.json done")

    # ================================================================
    # BUILD SLEEP JSON
    # ================================================================
    print("Building sleep.json ...")

    asleep_types = [
        'HKCategoryValueSleepAnalysisAsleepCore',
        'HKCategoryValueSleepAnalysisAsleepDeep',
        'HKCategoryValueSleepAnalysisAsleepREM',
        'HKCategoryValueSleepAnalysisAsleepUnspecified',
    ]

    # Group by night (using end_date)
    nightly = defaultdict(lambda: {'core': 0, 'deep': 0, 'rem': 0, 'total': 0, 'bedtime': None, 'wakeTime': None})
    for r in sleep_records:
        if r['value'] not in asleep_types:
            continue
        night = r['end_date']
        dur = r['duration_hrs']
        if dur <= 0 or dur > 12:
            continue
        nightly[night]['total'] += dur
        if 'Core' in r['value']:
            nightly[night]['core'] += dur
        elif 'Deep' in r['value']:
            nightly[night]['deep'] += dur
        elif 'REM' in r['value']:
            nightly[night]['rem'] += dur

        # Track bedtime (earliest start after 20:00) and wake time (latest end)
        start_dt = parse_date(r['start'])
        end_dt = parse_date(r['end'])
        if start_dt:
            bt_hour = start_dt.hour + start_dt.minute / 60
            # Skip records starting between 12:00-20:00 (likely naps)
            if bt_hour < 12:
                bt_hour += 24  # e.g., 3am → 27
            if bt_hour >= 20:  # only count sleep starting after 8pm
                if nightly[night]['bedtime'] is None or bt_hour < nightly[night]['bedtime']:
                    nightly[night]['bedtime'] = bt_hour
        if end_dt:
            wt_hour = end_dt.hour + end_dt.minute / 60
            if nightly[night]['wakeTime'] is None or wt_hour > nightly[night]['wakeTime']:
                nightly[night]['wakeTime'] = wt_hour

    # Filter valid nights
    nightly_list = []
    for date, data in sorted(nightly.items()):
        if 2 < data['total'] < 16:
            nightly_list.append({
                'date': date,
                'total': round(data['total'], 2),
                'core': round(data['core'], 2),
                'deep': round(data['deep'], 2),
                'rem': round(data['rem'], 2),
                'bedtime': round(data['bedtime'], 2) if data['bedtime'] else None,
                'wakeTime': round(data['wakeTime'], 2) if data['wakeTime'] else None,
            })

    # Sleep monthly
    sleep_monthly = defaultdict(list)
    for n in nightly_list:
        m = n['date'][:7]
        sleep_monthly[m].append(n)

    sleep_monthly_list = []
    for m, nights in sorted(sleep_monthly.items()):
        totals = [n['total'] for n in nights]
        deeps = [n['deep'] for n in nights if n['deep'] > 0]
        rems = [n['rem'] for n in nights if n['rem'] > 0]
        beds = [n['bedtime'] for n in nights if n['bedtime']]
        sleep_monthly_list.append({
            'month': m,
            'avgTotal': round(sum(totals)/len(totals), 2),
            'avgDeep': round(sum(deeps)/len(deeps), 2) if deeps else 0,
            'avgREM': round(sum(rems)/len(rems), 2) if rems else 0,
            'avgBedtime': round(sum(beds)/len(beds), 2) if beds else 0,
            'count': len(nights),
        })

    # Bedtime heatmap: [dayOfWeek][hourBucket] → count
    bedtime_heatmap = []
    for n in nightly_list:
        if n['bedtime'] is None:
            continue
        dt = datetime.strptime(n['date'], '%Y-%m-%d')
        dow = dt.weekday()
        hour_bucket = int(n['bedtime']) % 24
        bedtime_heatmap.append({'dow': dow, 'hour': hour_bucket, 'date': n['date']})

    # Aggregate heatmap
    heatmap_counts = defaultdict(int)
    for h in bedtime_heatmap:
        heatmap_counts[(h['dow'], h['hour'])] += 1
    heatmap_data = [{'dow': k[0], 'hour': k[1], 'count': v} for k, v in heatmap_counts.items()]

    # Sleep stats
    all_totals = [n['total'] for n in nightly_list]
    all_deeps = [n['deep'] for n in nightly_list if n['deep'] > 0]
    all_rems = [n['rem'] for n in nightly_list if n['rem'] > 0]
    all_beds = [n['bedtime'] for n in nightly_list if n['bedtime']]

    sleep_stats = {
        'avgTotal': round(sum(all_totals)/len(all_totals), 2) if all_totals else 0,
        'avgDeep': round(sum(all_deeps)/len(all_deeps), 2) if all_deeps else 0,
        'avgREM': round(sum(all_rems)/len(all_rems), 2) if all_rems else 0,
        'avgBedtime': round(sum(all_beds)/len(all_beds), 2) if all_beds else 0,
        'below6hPct': round(sum(1 for t in all_totals if t < 6)/len(all_totals)*100, 1) if all_totals else 0,
        'below7hPct': round(sum(1 for t in all_totals if t < 7)/len(all_totals)*100, 1) if all_totals else 0,
        'above8hPct': round(sum(1 for t in all_totals if t >= 8)/len(all_totals)*100, 1) if all_totals else 0,
        'totalNights': len(nightly_list),
        'deepPct': round(sum(all_deeps)/len(all_deeps) / (sum(all_totals)/len(all_totals)) * 100, 1) if all_deeps and all_totals else 0,
        'remPct': round(sum(all_rems)/len(all_rems) / (sum(all_totals)/len(all_totals)) * 100, 1) if all_rems and all_totals else 0,
    }

    # Breathing disturbances and wrist temperature
    sleep_breathing_sorted = sorted(sleep_breathing, key=lambda x: x['date'])
    sleep_temp_sorted = sorted(sleep_temp, key=lambda x: x['date'])

    sleep_data = {
        'nightly': nightly_list,
        'monthly': sleep_monthly_list,
        'heatmap': heatmap_data,
        'stats': sleep_stats,
        'breathingDisturbances': sleep_breathing_sorted,
        'wristTemperature': sleep_temp_sorted,
    }

    # Sleep debt computation
    sleep_debt_target = CONFIG.get('sleepDebt', {}).get('targetHours', 7.5)
    sleep_debt_max_days = CONFIG.get('sleepDebt', {}).get('maxDebtDays', 14)
    sleep_debt_data = []
    cumulative_debt = 0
    for n in nightly_list[-(sleep_debt_max_days * 4):]:
        deficit = sleep_debt_target - n['total']
        cumulative_debt = max(0, cumulative_debt + deficit)
        sleep_debt_data.append({
            'date': n['date'],
            'duration': round(n['total'], 2),
            'target': sleep_debt_target,
            'deficit': round(deficit, 2),
            'cumulativeDebt': round(cumulative_debt, 2),
        })
    sleep_data['sleepDebt'] = sleep_debt_data

    with open(os.path.join(out_dir, 'sleep.json'), 'w') as f:
        json.dump(sleep_data, f)
    print("  sleep.json done")

    # ================================================================
    # BUILD ACTIVITY JSON
    # ================================================================
    print("Building activity.json ...")

    # Daily steps
    steps_list = sorted([{'date': d, 'value': round(v)} for d, v in steps_daily.items() if v > 100], key=lambda x: x['date'])

    # Daily active energy
    energy_list = sorted([{'date': d, 'value': round(v)} for d, v in active_energy_daily.items() if v > 50], key=lambda x: x['date'])

    # Steps monthly
    steps_monthly = defaultdict(list)
    for s in steps_list:
        m = s['date'][:7]
        steps_monthly[m].append(s['value'])
    steps_monthly_list = sorted([{
        'month': m,
        'mean': round(sum(vs)/len(vs)),
        'median': round(sorted(vs)[len(vs)//2]),
    } for m, vs in steps_monthly.items()], key=lambda x: x['month'])

    # Steps stats
    step_values = [s['value'] for s in steps_list]
    steps_stats = {
        'mean': round(sum(step_values)/len(step_values)) if step_values else 0,
        'median': round(sorted(step_values)[len(step_values)//2]) if step_values else 0,
        'below5000pct': round(sum(1 for v in step_values if v < 5000)/len(step_values)*100, 1) if step_values else 0,
        'above10000pct': round(sum(1 for v in step_values if v >= 10000)/len(step_values)*100, 1) if step_values else 0,
    }

    # Workouts - already collected
    workouts_sorted = sorted(workouts, key=lambda x: x.get('date', ''))

    # Workout monthly frequency
    wo_monthly = defaultdict(int)
    for w in workouts_sorted:
        m = w.get('date', '')[:7]
        if m:
            wo_monthly[m] += 1
    wo_monthly_list = sorted([{'month': m, 'count': c} for m, c in wo_monthly.items()], key=lambda x: x['month'])

    # Swimming specific
    swim_workouts = [w for w in workouts_sorted if w['type'] == 'Swimming']

    # Activity summaries
    act_list = sorted(activity_summaries, key=lambda x: x['date'])

    # Body composition
    body_mass_sorted = sorted(body_mass, key=lambda x: x['date'])
    body_fat_sorted = sorted(body_fat, key=lambda x: x['date'])

    # Exercise time daily
    exercise_list = sorted([{'date': d, 'value': round(v)} for d, v in exercise_time_daily.items() if v > 0], key=lambda x: x['date'])

    activity_data = {
        'steps': {'daily': steps_list, 'monthly': steps_monthly_list, 'stats': steps_stats},
        'energy': {'daily': energy_list},
        'workouts': workouts_sorted,
        'swimWorkouts': swim_workouts,
        'workoutMonthly': wo_monthly_list,
        'activitySummary': act_list,
        'bodyMass': body_mass_sorted,
        'bodyFat': body_fat_sorted,
        'exerciseTime': {'daily': exercise_list},
    }

    # Walking metrics (monthly averages)
    def monthly_avg(daily_dict):
        monthly = defaultdict(list)
        for d, vals in sorted(daily_dict.items()):
            m = d[:7]
            if isinstance(vals, list):
                monthly[m].extend(vals)
            else:
                monthly[m].append(vals)
        return sorted([{'month': m, 'mean': round(sum(vs)/len(vs), 2)} for m, vs in monthly.items()], key=lambda x: x['month'])

    activity_data['walkingSpeed'] = {'monthly': monthly_avg(walking_speed_daily)}
    activity_data['walkingStepLength'] = {'monthly': monthly_avg(walking_step_length_daily)}
    activity_data['walkingAsymmetry'] = {'monthly': monthly_avg(walking_asymmetry_daily)}
    activity_data['walkingSteadiness'] = sorted(walking_steadiness, key=lambda x: x['date'])

    # Daylight
    daylight_list = sorted([{'date': d, 'value': round(v)} for d, v in daylight_daily.items() if v > 0], key=lambda x: x['date'])
    activity_data['daylight'] = {'daily': daylight_list, 'monthly': monthly_avg(daylight_daily)}

    # Distance
    distance_list = sorted([{'date': d, 'value': round(v, 2)} for d, v in distance_daily.items() if v > 0], key=lambda x: x['date'])
    activity_data['distance'] = {'daily': distance_list}

    # Flights
    flights_list = sorted([{'date': d, 'value': round(v)} for d, v in flights_daily.items() if v > 0], key=lambda x: x['date'])
    activity_data['flights'] = {'daily': flights_list}

    # Basal energy
    basal_list = sorted([{'date': d, 'value': round(v)} for d, v in basal_energy_daily.items() if v > 50], key=lambda x: x['date'])
    activity_data['basalEnergy'] = {'daily': basal_list}

    # Headphone exposure
    headphone_monthly = defaultdict(list)
    for d, vals in headphone_exposure_daily.items():
        m = d[:7]
        headphone_monthly[m].extend(vals)
    headphone_monthly_list = sorted([{
        'month': m, 'mean': round(sum(vs)/len(vs), 1), 'max': round(max(vs), 1),
    } for m, vs in headphone_monthly.items()], key=lambda x: x['month'])
    activity_data['headphoneExposure'] = {'monthly': headphone_monthly_list}

    # Cycling distance
    cycling_list = sorted([{'date': d, 'value': round(v, 2)} for d, v in cycling_distance_daily.items() if v > 0], key=lambda x: x['date'])
    activity_data['cyclingDistance'] = {'daily': cycling_list}

    # Six minute walk test
    activity_data['sixMinWalk'] = sorted(six_min_walk, key=lambda x: x['date'])

    # Parse Arboleaf body scale data if provided
    arboleaf_data = []
    if arboleaf_path and os.path.exists(arboleaf_path):
        print(f"Parsing Arboleaf data: {arboleaf_path}")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(arboleaf_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                entry = dict(zip(headers, row))
                date_val = entry.get('测量时间', '')
                if not date_val:
                    continue
                # Parse date: "04/10/2026 13:46:43" format
                try:
                    from datetime import datetime as dt_parse
                    dt = dt_parse.strptime(str(date_val), '%m/%d/%Y %H:%M:%S')
                    date_str = dt.strftime('%Y-%m-%d')
                except (ValueError, TypeError):
                    continue

                def safe_float(v):
                    try:
                        f = float(v)
                        return f if f > 0 else None
                    except (ValueError, TypeError):
                        return None

                record = {'date': date_str}
                field_map = {
                    '体重(kg)': 'weight',
                    '脂肪率(%)': 'bodyFat',
                    'BMI': 'bmi',
                    '骨骼肌率(%)': 'skeletalMuscle',
                    '肌肉量(kg)': 'muscleMass',
                    '蛋白质(%)': 'protein',
                    '基础代谢量(kcal)': 'bmr',
                    '去脂体重(kg)': 'leanMass',
                    '皮下脂肪率(%)': 'subcutaneousFat',
                    '内脏脂肪': 'visceralFat',
                    '体水份(%)': 'bodyWater',
                    '骨量(kg)': 'boneMass',
                }
                for cn_key, en_key in field_map.items():
                    val = safe_float(entry.get(cn_key))
                    if val is not None:
                        record[en_key] = round(val, 2)

                if len(record) > 1:  # has at least one metric besides date
                    arboleaf_data.append(record)

            wb.close()
            arboleaf_data.sort(key=lambda x: x['date'])
            print(f"  Arboleaf: {len(arboleaf_data)} body composition records")
        except ImportError:
            print("  WARNING: openpyxl not installed. Run: pip install openpyxl")
            print("  Arboleaf body composition data will be skipped.")
        except Exception as e:
            print(f"  WARNING: Failed to parse Arboleaf data: {e}")
            traceback.print_exc()

    if arboleaf_data:
        activity_data['arboleaf'] = arboleaf_data

    with open(os.path.join(out_dir, 'activity.json'), 'w') as f:
        json.dump(activity_data, f)
    print("  activity.json done")

    # ================================================================
    # BUILD OVERVIEW JSON
    # ================================================================
    print("Building overview.json ...")

    # Determine age
    dob = user_info.get('dob', '1990-01-01')
    try:
        dob_dt = datetime.strptime(dob, '%Y-%m-%d')
        age = (datetime.now() - dob_dt).days // 365
    except (ValueError, TypeError):
        age = 28

    # Anomaly detection
    anomalies = []
    rhr_mean = sum(rhr_values) / len(rhr_values) if rhr_values else 54
    rhr_std = (sum((v - rhr_mean)**2 for v in rhr_values) / len(rhr_values))**0.5 if rhr_values else 4
    for r in rhr_list:
        if r['value'] > rhr_mean + 2 * rhr_std:
            anomalies.append({'date': r['date'], 'type': 'rhr_high', 'value': r['value'], 'label': f"RHR {r['value']:.0f} bpm"})

    # HRV anomalies
    if hrv_daily_list:
        hrv_vals = [r['value'] for r in hrv_daily_list]
        hrv_m = sum(hrv_vals) / len(hrv_vals)
        hrv_s = (sum((v - hrv_m) ** 2 for v in hrv_vals) / len(hrv_vals)) ** 0.5
        for r in hrv_daily_list:
            if r['value'] < hrv_m - 2 * hrv_s:
                anomalies.append({'date': r['date'], 'type': 'hrv_low', 'value': r['value'], 'label': f"HRV {r['value']:.0f} ms"})

    # Sleep anomalies
    for n in nightly_list:
        if n['total'] < 4:
            anomalies.append({'date': n['date'], 'type': 'sleep_short', 'value': n['total'], 'label': f"Sleep {n['total']:.1f}h"})
        if n.get('bedtime') and n['bedtime'] > 28:
            anomalies.append({'date': n['date'], 'type': 'bedtime_late', 'value': n['bedtime'], 'label': "Bed after 4AM"})

    # Step anomalies
    if step_values:
        step_mean = sum(step_values) / len(step_values)
        if step_mean > 6000:
            for s in steps_list:
                if s['value'] < 2000:
                    anomalies.append({'date': s['date'], 'type': 'steps_low', 'value': s['value'], 'label': f"Steps {s['value']}"})

    # SpO2 anomalies
    for r in spo2_records:
        if r['value'] < 92:
            anomalies.append({'date': r['date'], 'type': 'spo2_low', 'value': round(r['value'], 1), 'label': f"SpO2 {r['value']:.0f}%"})

    # Compute health score
    health_score = compute_health_score(
        rhr_list, hrv_daily_map, nightly_list, steps_daily,
        workouts_sorted, body_mass_sorted, age,
        daylight_daily=daylight_daily,
    )

    # Compute longevity score
    longevity_score = compute_longevity_score(
        vo2_list, rhr_list, hrv_records, hrv_daily_map, nightly_list,
        steps_daily, workouts_sorted, body_mass_sorted, body_fat_sorted,
        spo2_records, exercise_time_daily, age,
        'male' if 'Male' in user_info.get('sex', '') else 'female',
        daylight_daily=daylight_daily,
    )

    # Compute baselines
    baselines = compute_baselines(rhr_list, hrv_daily_map, nightly_list, steps_daily)

    # Compute correlations
    correlations = compute_correlations(
        rhr_list, hrv_daily_map, nightly_list, steps_daily,
        workouts_sorted, daylight_daily
    )

    # Compute risks
    risks = compute_risks(
        rhr_list=rhr_list,
        hrv_records=hrv_records,
        vo2_list=vo2_list,
        nightly_list=nightly_list,
        steps_daily=steps_daily,
        body_mass=body_mass_sorted,
        body_fat=body_fat_sorted,
        spo2_records=spo2_records,
        sleep_breathing=sleep_breathing,
        workouts=workouts_sorted,
        exercise_time_daily=exercise_time_daily,
        all_beds=all_beds,
        age=age,
        daylight_daily=daylight_daily,
        headphone_exposure_daily=headphone_exposure_daily,
        walking_asymmetry_daily=walking_asymmetry_daily,
        arboleaf_data=arboleaf_data,
    )

    # Compute trends
    trends = compute_trends(rhr_list, hrv_daily_map, nightly_list, steps_daily, CONFIG)

    # Compute heart rate cycles
    hr_cycles = compute_hr_cycles(rhr_list, hrv_daily_map, workouts_sorted)
    if hr_cycles['avgCycleLength'] > 0:
        print(f"  HR cycles: avg {hr_cycles['avgCycleLength']} days, {len(hr_cycles['troughs'])} recovery troughs detected")
    if trends:
        print(f"  Trend alerts: {len(trends)} detected")

    llm_provider = CONFIG.get('llm', {}).get('provider', 'claude')
    print(f"Generating AI weekly report (provider: {llm_provider})...")
    weekly_report = generate_weekly_report(
        health_score, longevity_score, risks, baselines, correlations,
        rhr_list, hrv_daily_map, nightly_list, steps_daily,
        daylight_daily, workouts_sorted, arboleaf_data
    )
    if weekly_report.get('error'):
        print(f"  Weekly report: {weekly_report['error']}")
    else:
        print(f"  Weekly report generated ({weekly_report['generatedAt']})")

    overview = {
        'user': {
            'dob': dob,
            'age': age,
            'sex': 'male' if 'Male' in user_info.get('sex', '') else 'female',
            'currentWeight': body_mass_sorted[-1]['value'] if body_mass_sorted else 0,
            'currentBodyFat': body_fat_sorted[-1]['value'] if body_fat_sorted else 0,
            'latestWeightDate': body_mass_sorted[-1]['date'] if body_mass_sorted else '',
        },
        'risks': risks,
        'goals': compute_goals(
            rhr_stats=rhr_stats,
            hrv_stats=hrv_stats,
            steps_stats=steps_list,
            body_mass=body_mass_sorted,
            nightly_list=nightly_list,
            workouts=workouts_sorted,
            all_beds=all_beds,
        ),
        'anomalies': anomalies,
        'dataRange': {
            'start': rhr_list[0]['date'] if rhr_list else '',
            'end': rhr_list[-1]['date'] if rhr_list else '',
        },
        'healthScore': health_score,
        'longevityScore': longevity_score,
        'baselines': baselines,
        'correlations': correlations,
        'weeklyReport': weekly_report,
        'trends': trends,
        'hrCycles': hr_cycles,
        'sources': {name: {'total': sum(types.values()), 'types': dict(types)}
                    for name, types in sorted(source_stats.items(), key=lambda x: -sum(x[1].values()))},
    }

    with open(os.path.join(out_dir, 'overview.json'), 'w') as f:
        json.dump(overview, f)
    print("  overview.json done")

    # ================================================================
    # BUILD ECG JSON
    # ================================================================
    print("Building ecg.json ...")
    ecg_dir = os.path.join(export_dir, 'electrocardiograms')
    ecg_records = []
    if os.path.isdir(ecg_dir):
        import csv as csv_module
        for fname in sorted(os.listdir(ecg_dir)):
            if not fname.endswith('.csv'):
                continue
            fpath = os.path.join(ecg_dir, fname)
            meta = {}
            samples = []
            with open(fpath, 'r') as ef:
                reader = csv_module.reader(ef)
                for row in reader:
                    if len(row) == 2 and row[0] and not row[0].replace('.','').replace('-','').lstrip('-').isdigit():
                        meta[row[0].strip()] = row[1].strip()
                    elif len(row) == 1:
                        try:
                            samples.append(float(row[0]))
                        except (ValueError, TypeError):
                            pass

            # Downsample to ~256 points for display (from ~15000 at 512Hz)
            if samples:
                step = max(1, len(samples) // 256)
                downsampled = [round(samples[i]) for i in range(0, len(samples), step)][:256]
            else:
                downsampled = []

            ecg_records.append({
                'filename': fname,
                'date': meta.get('Recorded Date', fname.replace('ecg_','').replace('.csv','')),
                'classification': meta.get('Classification', 'Unknown'),
                'sampleRate': meta.get('Sample Rate', '512 hertz'),
                'samples': downsampled,
            })

    ecg_data = {'records': ecg_records}
    with open(os.path.join(out_dir, 'ecg.json'), 'w') as f:
        json.dump(ecg_data, f)
    print(f"  ecg.json done ({len(ecg_records)} recordings)")

    # Clean up temp dir from ZIP extraction
    if _temp_dir and os.path.exists(_temp_dir):
        import shutil
        shutil.rmtree(_temp_dir, ignore_errors=True)
        print(f"  Cleaned up temp dir: {_temp_dir}")

    print("\nAll data files generated successfully!")
    print(f"Output: {out_dir}/")
    print("\nNew fields summary:")
    print(f"  respiratory: {len(resp_monthly_list)} months")
    print(f"  spo2 hourly: {len(spo2_hourly)} hours")
    print(f"  breathing disturbances: {len(sleep_breathing_sorted)} records")
    print(f"  wrist temperature: {len(sleep_temp_sorted)} records")
    print(f"  exercise time: {len(exercise_list)} days")
    print(f"  health score: {len(health_score['daily'])} days, latest={health_score['latest']}")
    print(f"  longevity score: {longevity_score['score']}, {len(longevity_score['components'])} components")
    print(f"  baselines: {len(baselines)} metrics")
    print(f"  daylight: {len(daylight_list)} days")
    print(f"  walking metrics: speed {len(walking_speed_daily)}, step length {len(walking_step_length_daily)}, asymmetry {len(walking_asymmetry_daily)} days")
    print(f"  headphone exposure: {len(headphone_monthly_list)} months")
    print(f"  correlations: {len(correlations['pairs'])} significant pairs")
    print(f"  weekly report: {'generated' if weekly_report.get('en') else weekly_report.get('error', 'skipped')}")

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Process Apple Health export data')
    parser.add_argument('export_dir', help='Path to Apple Health export directory')
    parser.add_argument('--arboleaf', help='Path to Arboleaf body scale XLSX export', default=None)
    args = parser.parse_args()
    main(args.export_dir, arboleaf_path=args.arboleaf)
